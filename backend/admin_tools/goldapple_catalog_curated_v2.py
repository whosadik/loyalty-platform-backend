from __future__ import annotations

import csv
import json
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook

from admin_tools.goldapple_catalog import _load_workbook_rows, build_normalized_product_payload, json_default
from roadmap_app.fragrance_slots import slot_of_fragrance


CURATED_V2_CANONICAL_TYPES: dict[str, list[str]] = {
    "haircare": ["shampoo", "conditioner", "hair_mask", "hair_oil", "scalp_serum", "leave_in"],
    "skincare": ["cleanser", "serum", "moisturizer", "spf", "toner", "mask", "eye_cream", "essence"],
    "makeup": ["foundation", "mascara", "blush", "lipstick", "eyeshadow", "primer", "setting_spray"],
    "fragrance": ["edp", "edt", "body_mist"],
}

PRIORITY_COVERAGE_TYPES: tuple[tuple[str, str], ...] = (
    ("haircare", "leave_in"),
    ("skincare", "eye_cream"),
    ("skincare", "essence"),
    ("makeup", "primer"),
    ("makeup", "setting_spray"),
    ("haircare", "hair_mask"),
    ("haircare", "scalp_serum"),
)

SERIALIZED_FIELDS = ("concerns", "attrs", "actives", "flags", "supported_skin_types", "raw_meta")
SOURCE_FIELDS = ("raw_type", "type_title", "name", "url_slug")
SOURCE_WEIGHTS = {"raw_type": 6, "type_title": 6, "name": 3, "url_slug": 2}


@dataclass(frozen=True)
class MatchRule:
    category: str
    product_type: str
    patterns: tuple[str, ...]


@dataclass(frozen=True)
class RowOverride:
    status: str
    reason: str
    category: str | None = None
    product_type: str | None = None
    brand: str | None = None
    attrs: dict[str, Any] | None = None
    concerns: list[str] | None = None
    actives: list[str] | None = None
    flags: list[str] | None = None
    supported_skin_types: list[str] | None = None
    strength: str | None = None
    source_url: str | None = None


@dataclass(frozen=True)
class AddedProduct:
    key: str
    name: str
    brand: str
    price: str
    category: str
    product_type: str
    currency: str = "KZT"
    in_stock: bool = True
    description: str = ""
    application_text: str = ""
    ingredients_inci: str = ""
    attrs: dict[str, Any] = field(default_factory=dict)
    concerns: list[str] = field(default_factory=list)
    actives: list[str] = field(default_factory=list)
    flags: list[str] = field(default_factory=list)
    supported_skin_types: list[str] = field(default_factory=list)
    strength: str = ""
    image_url: str = ""
    image_urls: list[str] = field(default_factory=list)
    volume_raw: str = ""
    raw_type: str = ""
    raw_meta_extra: dict[str, Any] = field(default_factory=dict)
    source_urls: tuple[str, ...] = field(default_factory=tuple)


MATCH_RULES: tuple[MatchRule, ...] = (
    MatchRule("fragrance", "edp", (r"парфюмерн\w+\s+вода", r"eau de parfum", r"\bedp\b")),
    MatchRule("fragrance", "edt", (r"туалетн\w+\s+вода", r"eau de toilette", r"\bedt\b")),
    MatchRule("fragrance", "body_mist", (r"body mist", r"hair mist", r"mist for body", r"мист для тела")),
    MatchRule("haircare", "leave_in", (r"leave[- ]?in", r"несмыва", r"leave in conditioner", r"leave in hair treatment")),
    MatchRule("haircare", "hair_mask", (r"маск\w*\s+для\s+волос", r"\bhair mask\b")),
    MatchRule("haircare", "hair_oil", (r"масл\w*\s+для\s+волос", r"\bhair oil\b")),
    MatchRule("haircare", "scalp_serum", (r"(сыворотк\w*|serum).*(кож\w+\s+голов|scalp)", r"scalp serum")),
    MatchRule("haircare", "conditioner", (r"кондиционер", r"бальзам\w*\s+для\s+волос", r"\bconditioner\b")),
    MatchRule("haircare", "shampoo", (r"шампун", r"\bshampoo\b")),
    MatchRule("skincare", "spf", (r"\bspf\b", r"sunscreen", r"sun cream", r"sun stick", r"sun base")),
    MatchRule("skincare", "eye_cream", (r"крем\w*.*глаз", r"\beye cream\b")),
    MatchRule("skincare", "essence", (r"эссенц", r"\bessence\b")),
    MatchRule("skincare", "toner", (r"тонер", r"тоник", r"\btoner\b")),
    MatchRule("skincare", "mask", (r"маск\w*\s+для\s+лиц", r"sheet mask", r"face mask", r"mask pack")),
    MatchRule(
        "skincare",
        "cleanser",
        (
            r"пенк\w*.*лиц",
            r"гель\w*.*лиц",
            r"cleanser",
            r"cleansing",
            r"micellar",
            r"гидрофильн\w*\s+(масл|бальзам)",
            r"cleansing balm",
            r"cleansing foam",
        ),
    ),
    MatchRule("skincare", "serum", (r"сыворотк\w*.*лиц", r"\bserum\b", r"\bampoule\b", r"spray serum")),
    MatchRule(
        "skincare",
        "moisturizer",
        (r"крем\w*.*лиц", r"эмульси\w*.*лиц", r"флюид\w*.*лиц", r"gel cream", r"hydrator", r"emulsion"),
    ),
    MatchRule("makeup", "foundation", (r"тональн", r"\bfoundation\b", r"skin tint", r"bb boomer")),
    MatchRule("makeup", "mascara", (r"туш", r"\bmascara\b")),
    MatchRule("makeup", "blush", (r"румян", r"\bblush\b")),
    MatchRule("makeup", "lipstick", (r"губн\w+\s+помад", r"\blipstick\b", r"lip cloud", r"lip suit", r"lip glow")),
    MatchRule("makeup", "eyeshadow", (r"тен\w+\s+для\s+век", r"палетк\w+.*глаз", r"\beyeshadow\b", r"eye palette")),
    MatchRule("makeup", "primer", (r"праймер", r"\bprimer\b", r"база под макияж", r"face primer")),
    MatchRule(
        "makeup",
        "setting_spray",
        (
            r"setting spray",
            r"make-up fixing spray",
            r"makeup fixing spray",
            r"prep & set spray",
            r"фиксатор для макияжа",
            r"спрей-фиксатор",
        ),
    ),
)

OUTSIDE_SCOPE_PATTERNS: tuple[tuple[str, str], ...] = (
    ("body_or_bath", r"(для тела|body cream|body wash|body lotion|body balm|body yogurt|гель для душа|для душа|ванн|bath)"),
    ("hands_feet", r"(для рук|для ног|для стоп|hand cream|foot cream)"),
    ("household", r"(mouthwash|ополаскивател|стирк|laundry|влажн\w+\s+салфет)"),
    ("tools", r"(кист|массажер|ролик|brush|gift set|подарочн\w+\s+набор|набор)"),
    ("patches", r"(патч)"),
    ("face_mist", r"(face mist|мист\s+для\s+лица|спреи?\s+и\s+мисты\s+для\s+лица|спрей\s+для\s+лица)"),
    ("powder", r"(пудр|powder)"),
    ("brow_or_eye_pencil", r"(карандаш.*глаз|карандаш.*бров|brow|eyeliner|eye pencil|brow pencil)"),
    ("highlighter_concealer", r"(хайлайтер|highlighter|concealer)"),
    ("deodorant", r"(дезодорант|deodorant)"),
)

NEAR_ONTOLOGY_PATTERNS: tuple[tuple[str, str], ...] = (
    ("lip_balm", r"(бальзам.*губ|lip balm|lip gloss)"),
    ("face_oil", r"(масло.*лиц|face oil)"),
    ("scrub_or_polish", r"(скраб.*лиц|polish)"),
    ("hair_serum_non_scalp", r"(сыворотк\w*.*волос|hair serum)"),
)

PLACEHOLDER_DESCRIPTION_PATTERN = re.compile(
    r"\s+—\s+(?:serum|toner|edp|edt|hair_mask|scalp_serum)\s+in category\s+(?:skincare|haircare|fragrance)\.\s+designed for daily use\.",
    re.IGNORECASE,
)

ROW_OVERRIDES: dict[int, RowOverride] = {
    12: RowOverride(
        status="fixable",
        category="skincare",
        product_type="cleanser",
        attrs={"area": "face", "volume_ml": 120},
        concerns=["cleansing", "makeup_removal"],
        supported_skin_types=["all"],
        flags=["cruelty_free"],
        strength="low",
        source_url="https://goldapple.kz/19000183935-all-clean-balm",
        reason="Hydrophilic cleansing balm for face; current haircare conditioner label is wrong.",
    ),
    47: RowOverride(
        status="fixable",
        category="makeup",
        product_type="blush",
        brand="3INA",
        attrs={"finish": "natural", "coverage": "sheer", "shade_family": "berry", "tone_family": "fair", "waterproof": False, "effect": "radiance"},
        strength="low",
        source_url="https://goldapple.kz/19760313800-the-blush",
        reason="Product page and raw title both describe blush, not mascara.",
    ),
    60: RowOverride(
        status="fixable",
        category="makeup",
        product_type="lipstick",
        attrs={"finish": "metallic", "coverage": "medium", "effect": "hydrating", "shade_family": "copper", "tone_family": "warm", "waterproof": False},
        strength="low",
        source_url="https://goldapple.kz/19000316969-lipstick",
        reason="Raw type is lipstick; current haircare conditioner label is wrong.",
    ),
    81: RowOverride(
        status="fixable",
        category="makeup",
        product_type="lipstick",
        attrs={"finish": "matte", "coverage": "medium", "effect": "blurring", "waterproof": False},
        strength="medium",
        source_url="https://goldapple.kz/19000172616-velvet-blur",
        reason="Velvet Blur is a lip product; current haircare conditioner label is wrong.",
    ),
    100: RowOverride(
        status="fixable",
        category="skincare",
        product_type="cleanser",
        attrs={"area": "face", "volume_ml": 100},
        concerns=["cleansing", "oil_control"],
        supported_skin_types=["all"],
        flags=["cruelty_free"],
        strength="low",
        source_url="https://goldapple.kz/19000267965-matcha-clay-cleanser",
        reason="Product page describes a facial cleansing foam; moisturizer label is wrong.",
    ),
    136: RowOverride(
        status="fixable",
        category="skincare",
        product_type="eye_cream",
        attrs={"area": "eyes", "volume_ml": 15},
        concerns=["nourishing", "repair", "anti_aging"],
        supported_skin_types=["all"],
        flags=["cruelty_free"],
        strength="medium",
        source_url="https://goldapple.kz/19000018905-power-collagen-eye-cream",
        reason="Raw type and URL clearly indicate eye cream.",
    ),
    140: RowOverride(
        status="fixable",
        category="skincare",
        product_type="moisturizer",
        attrs={"area": "face", "volume_ml": 100},
        concerns=["hydration", "oil_control"],
        supported_skin_types=["combination", "normal", "oily"],
        strength="medium",
        source_url="https://goldapple.kz/19000043632-biome-matte-hydrator",
        reason="Face fluid / hydrator should stay in moisturizer class, not cleanser.",
    ),
}

FORCE_AMBIGUOUS_ROWS: dict[int, str] = {
    18: "Hair serum without explicit scalp semantics does not map safely to scalp_serum.",
    21: "Facial scrub/polish is adjacent to skincare but not a canonical roadmap action.",
    53: "Face oil does not map safely to a canonical skincare action.",
    110: "Lip balm should not be stretched into lipstick.",
    112: "Lip balm should not be stretched into lipstick.",
}

FORCE_REJECT_ROWS: dict[int, str] = {
    3: "Powder should not be stretched into cleanser.",
    4: "Fixing powder is outside the roadmap ontology.",
    5: "Eye patches are outside the roadmap ontology.",
    8: "Eye pencil is outside the roadmap ontology.",
    9: "Foot cream is outside the roadmap ontology.",
    10: "Face mist should not be stretched into fragrance body_mist.",
    14: "Body wash is outside the roadmap ontology.",
    15: "Face mist should not be stretched into fragrance body_mist.",
    16: "Face patches are outside the roadmap ontology.",
    24: "Brow pencil is outside the roadmap ontology.",
    25: "Body balm is outside the roadmap ontology.",
    36: "Body scrub is outside the roadmap ontology.",
    39: "Deodorant is outside the roadmap ontology.",
    43: "Brow marker is outside the roadmap ontology.",
    46: "Brow pencil is outside the roadmap ontology.",
    48: "Body wash is outside the roadmap ontology.",
    50: "Highlighter is outside the roadmap ontology.",
    56: "Body balm is outside the roadmap ontology.",
    67: "Powder is outside the roadmap ontology.",
    69: "Body cream/yogurt is outside the roadmap ontology.",
    70: "Eye stick is outside the roadmap ontology.",
    82: "Body scrub is outside the roadmap ontology.",
    84: "Deodorant spray is outside the roadmap ontology.",
    94: "Bath salt is outside the roadmap ontology.",
    101: "Eye and lip pencil is outside the roadmap ontology.",
    106: "Facial roller is outside the roadmap ontology.",
    108: "Patches are outside the roadmap ontology.",
    114: "Shower gel is outside the roadmap ontology.",
    122: "Mouthwash is outside the roadmap ontology.",
    124: "Wet wipes are outside the roadmap ontology.",
    128: "Laundry gel is outside the roadmap ontology.",
    130: "Gift/set soap bundle is outside the roadmap ontology.",
    131: "Makeup brush is outside the roadmap ontology.",
    134: "Makeup brush is outside the roadmap ontology.",
    139: "Gift set is outside the roadmap ontology.",
    141: "Eye pencil is outside the roadmap ontology.",
    143: "Powder is outside the roadmap ontology.",
    144: "Body cream is outside the roadmap ontology.",
}


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=json_default)


def _norm_text(value: Any) -> str:
    text = str(value or "").lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _url_slug(value: str) -> str:
    url = str(value or "").strip()
    if not url:
        return ""
    return _norm_text(urlparse(url).path.strip("/").split("/")[-1])


def _normalize_serialized_fields(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(payload)
    for field_name in SERIALIZED_FIELDS:
        raw_value = normalized.get(field_name)
        if field_name in {"attrs", "raw_meta"}:
            normalized[field_name] = json.loads(_json_dumps(raw_value or {}))
        else:
            normalized[field_name] = json.loads(_json_dumps(raw_value or []))
    return normalized


def _copy_with_updates(payload: dict[str, Any], **updates: Any) -> dict[str, Any]:
    patched = dict(payload)
    patched.update(updates)
    return _normalize_serialized_fields(patched)


def _source_texts(raw_row: dict[str, Any]) -> dict[str, str]:
    return {
        "raw_type": _norm_text(raw_row.get("product_type_raw")),
        "type_title": _norm_text(raw_row.get("Type")),
        "name": _norm_text(raw_row.get("name") or raw_row.get("Name4") or raw_row.get("Name")),
        "url_slug": _url_slug(str(raw_row.get("galinkbase_1x27n_109_URL") or "")),
    }


def _is_placeholder_row(raw_row: dict[str, Any]) -> bool:
    description = str(raw_row.get("description_text") or raw_row.get("Info1") or "")
    product_id = str(raw_row.get("id") or "")
    raw_type = str(raw_row.get("product_type_raw") or raw_row.get("Type") or "").strip()
    return bool(product_id.startswith("99") and not raw_type and PLACEHOLDER_DESCRIPTION_PATTERN.search(description))


def _score_matches(raw_row: dict[str, Any]) -> tuple[Counter[tuple[str, str]], list[str], list[str]]:
    texts = _source_texts(raw_row)
    joined = " ".join(texts.values())
    scores: Counter[tuple[str, str]] = Counter()
    for rule in MATCH_RULES:
        for field_name in SOURCE_FIELDS:
            field_text = texts[field_name]
            if field_text and any(re.search(pattern, field_text) for pattern in rule.patterns):
                scores[(rule.category, rule.product_type)] += SOURCE_WEIGHTS[field_name]
    outside = [name for name, pattern in OUTSIDE_SCOPE_PATTERNS if re.search(pattern, joined)]
    near = [name for name, pattern in NEAR_ONTOLOGY_PATTERNS if re.search(pattern, joined)]
    return scores, outside, near


ADDED_PRODUCTS: tuple[AddedProduct, ...] = (
    AddedProduct("leave_in_yellow_professional", "NUTRITIVE LEAVE-IN CONDITIONER", "Yellow Professional", "7302", "haircare", "leave_in", in_stock=False, description="Leave-in conditioner for dry hair.", raw_type="несмываемый кондиционер для волос", volume_raw="250 мл", source_urls=("https://goldapple.kz/kk/19000017686-nutritive-leave-in-conditioner",)),
    AddedProduct("leave_in_chi", "Leave-In Conditioner", "CHI", "15209", "haircare", "leave_in", description="Light leave-in conditioner for detangling, smoothing and heat styling support.", raw_type="несмываемый кондиционер для волос", volume_raw="118 мл", source_urls=("https://goldapple.kz/29230200006-leave-in-conditioner/?locale=kk",)),
    AddedProduct("leave_in_augustinus_bader", "THE LEAVE-IN HAIR TREATMENT", "Augustinus Bader", "56200", "haircare", "leave_in", description="Leave-in hair treatment for dry and brittle hair.", raw_type="leave-in hair treatment", volume_raw="100 мл", source_urls=("https://goldapple.kz/19000217177-the-leave-in-hair-treatment/",)),
    AddedProduct("eye_cream_dr_jart", "Ceramidin Eye Cream", "Dr.Jart+", "17239", "skincare", "eye_cream", in_stock=False, description="Moisturizing eye cream with barrier support focus.", attrs={"area": "eyes", "volume_ml": 15}, concerns=["hydration", "nourishing", "repair"], supported_skin_types=["sensitive"], raw_type="крем для глаз", volume_raw="15 мл", source_urls=("https://goldapple.kz/kk/19000031691-ceramidin-eye-cream",)),
    AddedProduct("eye_cream_maskoholic", "Lifting eye cream", "MASKOHOLIC", "5124", "skincare", "eye_cream", description="Eye cream with lifting and hydration focus.", attrs={"area": "eyes", "volume_ml": 20}, concerns=["hydration", "anti_aging", "lifting"], supported_skin_types=["all"], raw_type="крем для глаз", volume_raw="20 мл", source_urls=("https://goldapple.kz/kk/19000282710-lifting-eye-cream",)),
    AddedProduct("essence_rawquest", "MILK THISTLE BRIGHTENING SPA FIRST ESSENCE", "RAWQUEST", "16667", "skincare", "essence", description="First essence with milk thistle extract for hydration and tone-evening.", attrs={"area": "face", "volume_ml": 200}, concerns=["hydration", "nourishing"], supported_skin_types=["all"], raw_type="эссенция для лица", volume_raw="200 мл", source_urls=("https://goldapple.kz/19000036389-milk-thistle-brightening-spa-first-essence",)),
    AddedProduct("essence_isntree", "HYALURONIC ACID WATER ESSENCE", "IsNtree", "19370", "skincare", "essence", description="Hydrating hyaluronic acid essence for dehydrated skin.", attrs={"area": "face", "volume_ml": 50}, concerns=["hydration", "nourishing"], supported_skin_types=["dehydrated"], raw_type="эссенция для лица", volume_raw="50 мл", source_urls=("https://goldapple.kz/kk/19000006841-hyaluronic-acid-water-essence",)),
    AddedProduct("essence_miguhara", "Anti wrinkle First Essence origin", "MIGUHARA", "17189", "skincare", "essence", description="Anti-aging first essence with hydration support.", attrs={"area": "face", "volume_ml": 120}, concerns=["anti_aging", "hydration", "nourishing"], supported_skin_types=["all"], raw_type="эссенция для лица", volume_raw="120 мл", source_urls=("https://goldapple.kz/19000142135-anti-wrinkle-first-essence-origin",)),
    AddedProduct("primer_shikstudio", "FACE PRIMER", "SHIKstudio", "7175", "makeup", "primer", description="Face primer with SPF 15 for smoothing texture and extending wear.", attrs={"finish": "radiant", "effect": "pore_blur", "spf": 15, "waterproof": False}, raw_type="праймер для лица", volume_raw="30 мл", source_urls=("https://goldapple.kz/19000067183-face-primer/?locale=kk",)),
    AddedProduct("primer_holika", "Puri Pore No Sebum Primer Deep Pore", "Holika Holika", "3700", "makeup", "primer", description="Pore-focused face primer used as a makeup base.", attrs={"effect": "pore_blur"}, raw_type="праймер для лица", source_urls=("https://goldapple.kz/azija/makijazh/dlja-lica/prajmery?brandobjectid=37301",)),
    AddedProduct("primer_darling_prime_time", "Prime Time", "DARLING*", "9666", "makeup", "primer", description="Mattifying face primer used as a makeup base.", attrs={"finish": "matte", "effect": "oil_control"}, raw_type="матирующий праймер для лица", source_urls=("https://goldapple.kz/makijazh/lico?p=3",)),
    AddedProduct("setting_spray_frudia", "Re:proust Perfect Shield Make Up Setting Fixer", "Frudia", "9956", "makeup", "setting_spray", description="Hydrating makeup setting fixer.", attrs={"effect": "hydrating", "waterproof": False}, raw_type="увлажняющий спрей-фиксатор для макияжа", source_urls=("https://goldapple.kz/makijazh/lico/prajmery?brandobjectid=25270",)),
    AddedProduct("setting_spray_shikstudio", "Flawless makeup fix spray", "SHIKstudio", "7600", "makeup", "setting_spray", description="Makeup fixing spray for setting the finished look.", attrs={"effect": "fixing", "waterproof": False}, raw_type="спрей для фиксации макияжа", source_urls=("https://goldapple.kz/makijazh/lico/prajmery?brandobjectid=25270",)),
    AddedProduct("setting_spray_artdeco", "3 in 1 make-up fixing spray", "Artdeco", "5930", "makeup", "setting_spray", description="Multi-purpose makeup fixing spray.", attrs={"effect": "fixing", "waterproof": False}, raw_type="спрей для фиксации макияжа", source_urls=("https://goldapple.kz/makijazh/lico?p=3",)),
    AddedProduct("hair_mask_the_act", "hair mask", "The Act", "3646", "haircare", "hair_mask", description="Hair mask for dry hair with nourishing oils and proteins.", raw_type="маска для волос", volume_raw="150 мл", source_urls=("https://goldapple.kz/kk/19000036189-hair-mask",)),
    AddedProduct("hair_mask_allmasil_volume", "8 seconds volume hair mask", "ALLMASIL", "8670", "haircare", "hair_mask", description="Volume-focused hair mask for damaged and fine hair.", raw_type="маска для волос", volume_raw="100 мл", source_urls=("https://goldapple.kz/kk/19000274513-8-seconds-volume-hair-mask",)),
    AddedProduct("hair_mask_allmasil_repair", "8 seconds salon repair hair mask", "ALLMASIL", "15174", "haircare", "hair_mask", description="Repair hair mask for smoothness, shine and elasticity.", raw_type="маска для волос", volume_raw="200 мл", source_urls=("https://goldapple.kz/kk/19000274510-8-seconds-salon-repair-hair-mask",)),
    AddedProduct("scalp_serum_kerastase", "Nutri-Supplement Scalp Serum", "Kérastase", "28560", "haircare", "scalp_serum", description="Moisturizing scalp serum for dry scalps.", attrs={"scalp_type": "dry"}, raw_type="сыворотка для кожи головы", volume_raw="90 мл", source_urls=("https://www.kerastase.com/products/nutritive/scalp-serum?prescripted=hairdiag", "https://goldapple.kz/kk/f/professionalnaya-syvorotka-dlya-volos?locale=kk")),
    AddedProduct("scalp_serum_davines", "ENERGIZING Superactive", "Davines", "61182", "haircare", "scalp_serum", description="Scalp serum with density-support and anti-fall positioning.", raw_type="сыворотка для кожи головы", volume_raw="100 мл", source_urls=("https://worldes.davines.com/products/energizing-superactive", "https://goldapple.kz/kk/f/professionalnaya-syvorotka-dlya-volos?locale=kk")),
    AddedProduct("scalp_serum_serioxyl", "Serioxyl Advanced Denser Hair Density Activator Serum", "L'Oréal Professionnel", "30037", "haircare", "scalp_serum", description="Density activator serum applied directly on the scalp.", raw_type="сыворотка для кожи головы", volume_raw="90 мл", source_urls=("https://www.lorealprofessionnel.com/all-products/hair-care/serioxyl-advanced-denser-hair-density-activator-serum", "https://goldapple.kz/kk/f/professionalnaya-syvorotka-dlya-volos?locale=kk")),
)


def _apply_override(raw_row: dict[str, Any], payload: dict[str, Any], override: RowOverride) -> dict[str, Any]:
    updated = _copy_with_updates(
        payload,
        category=override.category or payload["category"],
        product_type=override.product_type or payload["product_type"],
        brand=override.brand or payload["brand"],
        attrs=override.attrs if override.attrs is not None else payload["attrs"],
        concerns=override.concerns if override.concerns is not None else payload["concerns"],
        actives=override.actives if override.actives is not None else payload["actives"],
        flags=override.flags if override.flags is not None else payload["flags"],
        supported_skin_types=override.supported_skin_types if override.supported_skin_types is not None else payload["supported_skin_types"],
        strength=override.strength if override.strength is not None else payload["strength"],
    )
    raw_meta = dict(updated.get("raw_meta") or {})
    raw_meta.update(
        {
            "curation_v2_status": override.status,
            "curation_v2_reason": override.reason,
            "original_category": payload["category"],
            "original_product_type": payload["product_type"],
            "source_url": override.source_url or raw_meta.get("raw_url") or raw_row.get("galinkbase_1x27n_109_URL"),
        }
    )
    updated["raw_meta"] = raw_meta
    return updated


def _build_added_payload(product: AddedProduct) -> dict[str, Any]:
    source_url = product.source_urls[0] if product.source_urls else ""
    match = re.search(r"/(\d+)-", source_url)
    source_product_id = match.group(1) if match else ""
    return _normalize_serialized_fields(
        {
            "source_product_id": source_product_id,
            "name": product.name,
            "brand": product.brand,
            "price": product.price,
            "currency": product.currency,
            "category": product.category,
            "product_type": product.product_type,
            "concerns": list(product.concerns),
            "attrs": dict(product.attrs),
            "actives": list(product.actives),
            "flags": list(product.flags),
            "supported_skin_types": list(product.supported_skin_types),
            "strength": product.strength,
            "in_stock": product.in_stock,
            "step": "",
            "image_url": product.image_url,
            "image_urls": list(product.image_urls),
            "description": product.description,
            "application_text": product.application_text,
            "ingredients_inci": product.ingredients_inci,
            "volume_raw": product.volume_raw,
            "raw_meta": {
                "catalog_source": "goldapple_curated_v2_web_addition",
                "curation_v2_status": "added",
                "curation_v2_reason": "Real product added from retailer/manufacturer sources to close canonical coverage gap.",
                "source_key": product.key,
                "source_url": source_url,
                "source_urls": list(product.source_urls),
                "raw_type": product.raw_type,
                **product.raw_meta_extra,
            },
        }
    )


def _coverage_for_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for category, product_types in CURATED_V2_CANONICAL_TYPES.items():
        category_rows = [row for row in rows if row["category"] == category]
        counts = {product_type: 0 for product_type in product_types}
        for row in category_rows:
            if row["product_type"] in counts:
                counts[row["product_type"]] += 1
        result[category] = {
            "total_rows": len(category_rows),
            "counts": counts,
            "missing": [product_type for product_type, count in counts.items() if count == 0],
            "under_target": [product_type for product_type, count in counts.items() if count < 2],
        }
    return result


def _fragrance_slot_distribution(rows: list[dict[str, Any]]) -> dict[str, int]:
    distribution: Counter[str] = Counter()
    for row in rows:
        if row["category"] != "fragrance":
            continue
        slot = slot_of_fragrance(row.get("attrs") or {}, row.get("raw_meta") or {})
        distribution[str(slot or "unknown")] += 1
    return dict(sorted(distribution.items()))


def _report_verdict(coverage_after: dict[str, Any]) -> dict[str, str]:
    unresolved = [
        (category, product_type)
        for category, payload in coverage_after.items()
        for product_type in payload["under_target"]
        if category != "fragrance"
    ]
    return {
        "safe_for_demo_catalog": "yes" if not unresolved else "no",
        "safe_for_runtime_catalog": "no",
        "safe_for_dataset_rebuild": "no",
    }


def build_curated_catalog_v2(xlsx_path: str) -> dict[str, Any]:
    workbook = _load_workbook_rows(xlsx_path)
    source_rows = workbook["rows"]
    normalized_rows = [_normalize_serialized_fields(build_normalized_product_payload(row)) for row in source_rows]

    audit_rows: list[dict[str, Any]] = []
    changes: list[dict[str, Any]] = []
    unchanged_row_ids: list[int] = []
    ambiguous_row_ids: list[int] = []
    rejected_row_ids: list[int] = []
    importable_rows: list[dict[str, Any]] = []
    importable_before_additions: list[dict[str, Any]] = []
    status_counts = Counter()

    for raw_row, payload in zip(source_rows, normalized_rows):
        source_row = int(raw_row.get("__source_row__") or 0)
        original_category = payload["category"]
        original_product_type = payload["product_type"]
        scores, outside_reasons, near_reasons = _score_matches(raw_row)
        top_matches = scores.most_common()
        best_pair = top_matches[0][0] if top_matches else None
        best_score = top_matches[0][1] if top_matches else 0
        current_pair = (payload["category"], payload["product_type"])

        status = "confident_correct"
        reason = "Current normalized category/product_type is semantically supported by concise raw fields."
        curated_payload = payload
        source_url = str(raw_row.get("galinkbase_1x27n_109_URL") or payload.get("raw_meta", {}).get("raw_url") or "")
        change_action = "keep"

        if source_row in ROW_OVERRIDES:
            override = ROW_OVERRIDES[source_row]
            status = override.status
            reason = override.reason
            curated_payload = _apply_override(raw_row, payload, override)
            source_url = override.source_url or source_url
            change_action = "fix"
        elif source_row in FORCE_REJECT_ROWS:
            status = "reject"
            reason = FORCE_REJECT_ROWS[source_row]
            change_action = "exclude"
        elif source_row in FORCE_AMBIGUOUS_ROWS:
            status = "ambiguous"
            reason = FORCE_AMBIGUOUS_ROWS[source_row]
            change_action = "exclude"
        elif _is_placeholder_row(raw_row):
            status = "reject"
            reason = "Synthetic/generated placeholder row without verifiable product source."
            change_action = "exclude"
        elif outside_reasons:
            status = "reject"
            reason = "Outside roadmap ontology: " + ", ".join(sorted(outside_reasons))
            change_action = "exclude"
        elif current_pair == ("skincare", "spf") and scores.get(("skincare", "spf"), 0) >= 6:
            status = "confident_correct"
            reason = "SPF cues in concise raw fields outweigh the generic gel/cream format."
        elif best_pair and current_pair == best_pair and best_score >= 6:
            status = "confident_correct"
        elif best_pair and current_pair != best_pair and best_score >= 6:
            status = "fixable"
            reason = f"Strong raw/title signal points to {best_pair[0]}:{best_pair[1]} instead of {current_pair[0]}:{current_pair[1]}."
            curated_payload = _copy_with_updates(payload, category=best_pair[0], product_type=best_pair[1])
            raw_meta = dict(curated_payload.get("raw_meta") or {})
            raw_meta.update(
                {
                    "curation_v2_status": "fixable",
                    "curation_v2_reason": reason,
                    "original_category": original_category,
                    "original_product_type": original_product_type,
                    "source_url": source_url,
                }
            )
            curated_payload["raw_meta"] = raw_meta
            change_action = "fix"
        elif best_pair and current_pair == best_pair and best_score >= 3:
            status = "confident_correct"
            reason = "Current normalized category/product_type is consistent with concise raw signals."
        elif near_reasons:
            status = "ambiguous"
            reason = "Near ontology but not safely mappable: " + ", ".join(sorted(near_reasons))
            change_action = "exclude"
        elif payload["category"] in CURATED_V2_CANONICAL_TYPES and payload["product_type"] in CURATED_V2_CANONICAL_TYPES[payload["category"]]:
            status = "confident_correct"
            reason = "No contradictory concise signal found; kept unchanged to avoid rewriting semantically plausible rows."
        else:
            status = "reject"
            reason = "Cannot map safely to canonical roadmap ontology."
            change_action = "exclude"

        status_counts[status] += 1
        if status == "confident_correct":
            unchanged_row_ids.append(source_row)
            importable_before_additions.append(curated_payload)
            importable_rows.append(curated_payload)
        elif status == "fixable":
            changes.append(
                {
                    "source_row": source_row,
                    "name": payload["name"],
                    "brand": curated_payload["brand"],
                    "old_category": original_category,
                    "old_product_type": original_product_type,
                    "new_category": curated_payload["category"],
                    "new_product_type": curated_payload["product_type"],
                    "reason": reason,
                    "source_url": source_url,
                }
            )
            importable_before_additions.append(curated_payload)
            importable_rows.append(curated_payload)
        elif status == "ambiguous":
            ambiguous_row_ids.append(source_row)
        else:
            rejected_row_ids.append(source_row)

        audit_rows.append(
            {
                "source_row": source_row,
                "name": payload["name"],
                "brand": payload["brand"],
                "original_category": original_category,
                "original_product_type": original_product_type,
                "curation_status": status,
                "curation_reason": reason,
                "final_category": curated_payload["category"] if status in {"confident_correct", "fixable"} else "",
                "final_product_type": curated_payload["product_type"] if status in {"confident_correct", "fixable"} else "",
                "source_url": source_url,
                "top_matches": [f"{category}:{product_type}:{score}" for (category, product_type), score in top_matches[:3]],
                "raw_type": str(raw_row.get("product_type_raw") or raw_row.get("Type") or ""),
                "change_action": change_action,
            }
        )

    added_rows = [_build_added_payload(product) for product in ADDED_PRODUCTS]
    importable_rows.extend(added_rows)

    coverage_before = _coverage_for_rows(normalized_rows)
    coverage_curated_before_additions = _coverage_for_rows(importable_before_additions)
    coverage_after = _coverage_for_rows(importable_rows)

    return {
        "xlsx_path": str(Path(xlsx_path).resolve()),
        "selected_sheet": workbook["selected_sheet"],
        "source_rows_total": len(source_rows),
        "audit_rows": audit_rows,
        "status_counts": dict(status_counts),
        "unchanged_row_ids": unchanged_row_ids,
        "ambiguous_row_ids": ambiguous_row_ids,
        "rejected_row_ids": rejected_row_ids,
        "changes": changes,
        "added_rows": added_rows,
        "coverage_before": coverage_before,
        "coverage_curated_before_additions": coverage_curated_before_additions,
        "coverage_after": coverage_after,
        "slot_distribution_before": _fragrance_slot_distribution(importable_before_additions),
        "slot_distribution_after": _fragrance_slot_distribution(importable_rows),
        "verdict": _report_verdict(coverage_after),
        "importable_rows": importable_rows,
        "priority_coverage": {
            f"{category}:{product_type}": coverage_after[category]["counts"].get(product_type, 0)
            for category, product_type in PRIORITY_COVERAGE_TYPES
        },
        "added_product_urls": {
            row["name"]: list((row.get("raw_meta") or {}).get("source_urls") or []) for row in added_rows
        },
    }


def _write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            serialized = dict(row)
            for key, value in list(serialized.items()):
                if isinstance(value, (dict, list)):
                    serialized[key] = _json_dumps(value)
            writer.writerow(serialized)


def _sheet_rows_for_importable(rows: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    headers = ["source_product_id", "name", "brand", "price", "currency", "category", "product_type", "concerns", "attrs", "actives", "flags", "supported_skin_types", "strength", "in_stock", "step", "image_url", "image_urls", "description", "application_text", "ingredients_inci", "volume_raw", "raw_meta"]
    sheet_rows: list[list[Any]] = []
    for row in rows:
        sheet_rows.append(
            [
                row.get("source_product_id", ""),
                row.get("name", ""),
                row.get("brand", ""),
                row.get("price", ""),
                row.get("currency", ""),
                row.get("category", ""),
                row.get("product_type", ""),
                _json_dumps(row.get("concerns") or []),
                _json_dumps(row.get("attrs") or {}),
                _json_dumps(row.get("actives") or []),
                _json_dumps(row.get("flags") or []),
                _json_dumps(row.get("supported_skin_types") or []),
                row.get("strength", ""),
                bool(row.get("in_stock", True)),
                row.get("step", ""),
                row.get("image_url", ""),
                _json_dumps(row.get("image_urls") or []),
                row.get("description", ""),
                row.get("application_text", ""),
                row.get("ingredients_inci", ""),
                row.get("volume_raw", ""),
                _json_dumps(row.get("raw_meta") or {}),
            ]
        )
    return headers, sheet_rows


def write_curated_catalog_v2_artifacts(result: dict[str, Any], workbook_path: str, audit_md_path: str, audit_json_path: str, changes_csv_path: str, added_products_csv_path: str) -> dict[str, str]:
    workbook_path_obj = Path(workbook_path).resolve()
    audit_md_obj = Path(audit_md_path).resolve()
    audit_json_obj = Path(audit_json_path).resolve()
    changes_csv_obj = Path(changes_csv_path).resolve()
    added_csv_obj = Path(added_products_csv_path).resolve()
    for parent in {workbook_path_obj.parent, audit_md_obj.parent, audit_json_obj.parent, changes_csv_obj.parent, added_csv_obj.parent}:
        parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_curated = wb.active
    ws_curated.title = "curated_catalog"
    curated_headers, curated_rows = _sheet_rows_for_importable(result["importable_rows"])
    ws_curated.append(curated_headers)
    for row in curated_rows:
        ws_curated.append(row)

    audit_headers = ["source_row", "name", "brand", "original_category", "original_product_type", "curation_status", "curation_reason", "final_category", "final_product_type", "source_url", "top_matches", "raw_type", "change_action"]
    ws_audit = wb.create_sheet("audit_all_rows")
    ws_audit.append(audit_headers)
    for row in result["audit_rows"]:
        ws_audit.append([row.get(header, "") if header != "top_matches" else " | ".join(row.get("top_matches") or []) for header in audit_headers])

    change_headers = ["source_row", "name", "brand", "old_category", "old_product_type", "new_category", "new_product_type", "reason", "source_url"]
    ws_changes = wb.create_sheet("changes")
    ws_changes.append(change_headers)
    for row in result["changes"]:
        ws_changes.append([row.get(header, "") for header in change_headers])

    added_headers = ["name", "brand", "price", "category", "product_type", "in_stock", "source_urls", "raw_type", "volume_raw"]
    ws_added = wb.create_sheet("added_products")
    ws_added.append(added_headers)
    added_csv_rows: list[dict[str, Any]] = []
    for row in result["added_rows"]:
        added_entry = {
            "name": row.get("name", ""),
            "brand": row.get("brand", ""),
            "price": row.get("price", ""),
            "category": row.get("category", ""),
            "product_type": row.get("product_type", ""),
            "in_stock": row.get("in_stock", True),
            "source_urls": " | ".join((row.get("raw_meta") or {}).get("source_urls") or []),
            "raw_type": (row.get("raw_meta") or {}).get("raw_type", ""),
            "volume_raw": row.get("volume_raw", ""),
        }
        added_csv_rows.append(added_entry)
        ws_added.append([added_entry[h] for h in added_headers])
    wb.save(workbook_path_obj)

    _write_csv(changes_csv_obj, result["changes"], change_headers)
    _write_csv(added_csv_obj, added_csv_rows, added_headers)

    json_payload = {
        "xlsx_path": result["xlsx_path"],
        "selected_sheet": result["selected_sheet"],
        "source_rows_total": result["source_rows_total"],
        "status_counts": result["status_counts"],
        "fixed_rows": result["changes"],
        "added_rows_count": len(result["added_rows"]),
        "added_product_urls": result["added_product_urls"],
        "coverage_before": result["coverage_before"],
        "coverage_curated_before_additions": result["coverage_curated_before_additions"],
        "coverage_after": result["coverage_after"],
        "slot_distribution_before": result["slot_distribution_before"],
        "slot_distribution_after": result["slot_distribution_after"],
        "priority_coverage": result["priority_coverage"],
        "verdict": result["verdict"],
        "unchanged_row_ids": result["unchanged_row_ids"],
        "ambiguous_row_ids": result["ambiguous_row_ids"],
        "rejected_row_ids": result["rejected_row_ids"],
    }
    audit_json_obj.write_text(_json_dumps(json_payload), encoding="utf-8")

    coverage_lines = []
    for category, payload in result["coverage_after"].items():
        counts = ", ".join(f"{product_type}={count}" for product_type, count in payload["counts"].items())
        coverage_lines.append(f"- {category}: {counts}; missing={payload['missing'] or 'none'}; under_target={payload['under_target'] or 'none'}")
    source_url_lines = [f"- {name}: " + " | ".join(urls) for name, urls in sorted(result["added_product_urls"].items())]
    markdown = "\n".join(
        [
            "# Goldapple Catalog Curated v2 Audit",
            "",
            f"- Source file: `{result['xlsx_path']}`",
            f"- Selected sheet: `{result['selected_sheet']}`",
            f"- Correct and untouched: **{result['status_counts'].get('confident_correct', 0)}**",
            f"- Fixed rows: **{result['status_counts'].get('fixable', 0)}**",
            f"- Ambiguous rows excluded: **{result['status_counts'].get('ambiguous', 0)}**",
            f"- Rejected rows excluded: **{result['status_counts'].get('reject', 0)}**",
            f"- Added real products: **{len(result['added_rows'])}**",
            "",
            "## Coverage After Curation",
            *coverage_lines,
            "",
            "## Fragrance Slot Distribution",
            f"- Before additions: `{json.dumps(result['slot_distribution_before'], ensure_ascii=False, sort_keys=True)}`",
            f"- After additions: `{json.dumps(result['slot_distribution_after'], ensure_ascii=False, sort_keys=True)}`",
            "",
            "## Fixed Rows",
            *[
                f"- row {row['source_row']}: {row['old_category']}/{row['old_product_type']} -> {row['new_category']}/{row['new_product_type']} ({row['reason']})"
                for row in result["changes"]
            ],
            "",
            "## Added Product Sources",
            *source_url_lines,
            "",
            "## Verdict",
            f"- safe_for_demo_catalog: **{result['verdict']['safe_for_demo_catalog']}**",
            f"- safe_for_runtime_catalog: **{result['verdict']['safe_for_runtime_catalog']}**",
            f"- safe_for_dataset_rebuild: **{result['verdict']['safe_for_dataset_rebuild']}**",
        ]
    ) + "\n"
    audit_md_obj.write_text(markdown, encoding="utf-8")

    return {
        "workbook": str(workbook_path_obj),
        "audit_md": str(audit_md_obj),
        "audit_json": str(audit_json_obj),
        "changes_csv": str(changes_csv_obj),
        "added_products_csv": str(added_csv_obj),
    }
