from __future__ import annotations

import json
import hashlib
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from admin_tools.goldapple_catalog import json_default
from admin_tools.goldapple_catalog_curated_v2 import CURATED_V2_CANONICAL_TYPES
from roadmap_app.fragrance_slots import slot_of_fragrance


CURATED_WORKBOOK_SHEET = "curated_catalog"
LIST_JSON_FIELDS = {"concerns", "actives", "flags", "supported_skin_types", "image_urls"}
DICT_JSON_FIELDS = {"attrs", "raw_meta"}
ALL_JSON_FIELDS = LIST_JSON_FIELDS | DICT_JSON_FIELDS
REQUIRED_HEADERS = {
    "source_product_id",
    "name",
    "brand",
    "price",
    "currency",
    "category",
    "product_type",
    "attrs",
    "raw_meta",
    "in_stock",
}
CRITICAL_ROW_FIELDS = {"name", "category", "product_type", "price"}
FRAGRANCE_RETAIL_TYPES = tuple(CURATED_V2_CANONICAL_TYPES["fragrance"])
ROADMAP_ACTION_TYPES = {
    category: tuple(types)
    for category, types in CURATED_V2_CANONICAL_TYPES.items()
    if category != "fragrance"
}


@dataclass(frozen=True)
class WorkbookRowIssue:
    sheet_row: int
    field: str
    reason: str
    raw_value: Any = None
    severity: str = "blocking"

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet_row": int(self.sheet_row),
            "field": self.field,
            "reason": self.reason,
            "raw_value": self.raw_value,
            "severity": self.severity,
        }


def _json_load_cell(value: Any, *, field_name: str, sheet_row: int, issues: list[WorkbookRowIssue]) -> Any:
    if value in (None, ""):
        return [] if field_name in LIST_JSON_FIELDS else {}
    if isinstance(value, (list, dict)):
        return value
    if not isinstance(value, str):
        issues.append(
            WorkbookRowIssue(
                sheet_row=sheet_row,
                field=field_name,
                reason="JSON field must be string/list/dict.",
                raw_value=value,
            )
        )
        return [] if field_name in LIST_JSON_FIELDS else {}
    try:
        parsed = json.loads(value)
    except Exception as exc:
        issues.append(
            WorkbookRowIssue(
                sheet_row=sheet_row,
                field=field_name,
                reason=f"JSON parse error: {exc}",
                raw_value=value,
            )
        )
        return [] if field_name in LIST_JSON_FIELDS else {}
    if field_name in LIST_JSON_FIELDS and not isinstance(parsed, list):
        issues.append(
            WorkbookRowIssue(
                sheet_row=sheet_row,
                field=field_name,
                reason="Expected JSON list.",
                raw_value=value,
            )
        )
        return []
    if field_name in DICT_JSON_FIELDS and not isinstance(parsed, dict):
        issues.append(
            WorkbookRowIssue(
                sheet_row=sheet_row,
                field=field_name,
                reason="Expected JSON object.",
                raw_value=value,
            )
        )
        return {}
    return parsed


def _normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y"}


def _normalize_price(value: Any, *, sheet_row: int, issues: list[WorkbookRowIssue]) -> str:
    if value in (None, ""):
        issues.append(WorkbookRowIssue(sheet_row=sheet_row, field="price", reason="Missing price."))
        return ""
    try:
        normalized = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        issues.append(
            WorkbookRowIssue(
                sheet_row=sheet_row,
                field="price",
                reason="Price is not a valid decimal.",
                raw_value=value,
            )
        )
        return ""
    return format(normalized.quantize(Decimal("0.01")), "f")


def _generated_source_product_id(row: dict[str, Any], *, sheet_row: int) -> str:
    brand = str(row.get("brand") or "").strip().lower().replace(" ", "-")
    name = str(row.get("name") or "").strip().lower().replace(" ", "-")
    product_type = str(row.get("product_type") or "").strip().lower()
    key = "-".join(part for part in (brand, name, product_type) if part)
    key = "".join(ch for ch in key if ch.isalnum() or ch in {"-", "_"})
    if not key:
        key = f"row-{sheet_row}"
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]
    return f"cv2:{sheet_row}:{digest}"


def _nested_product_type_counts(rows: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    result: dict[str, dict[str, int]] = {}
    for category, allowed_types in CURATED_V2_CANONICAL_TYPES.items():
        counter = Counter()
        for row in rows:
            if row.get("category") != category:
                continue
            product_type = str(row.get("product_type") or "")
            counter[product_type] += 1
        result[category] = {ptype: int(counter.get(ptype, 0)) for ptype in allowed_types}
    return result


def coverage_from_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_category = Counter()
    by_product_type = Counter()
    in_stock = Counter()
    fragrance_slots = Counter()
    for row in rows:
        category = str(row.get("category") or "")
        product_type = str(row.get("product_type") or "")
        by_category[category] += 1
        by_product_type[f"{category}:{product_type}"] += 1
        in_stock["true" if bool(row.get("in_stock")) else "false"] += 1
        if category == "fragrance":
            fragrance_slots[slot_of_fragrance(row.get("attrs") or {}, raw_meta=row.get("raw_meta") or {})] += 1
    return {
        "products_total": len(rows),
        "by_category": dict(by_category),
        "by_product_type": dict(by_product_type),
        "canonical_coverage": _nested_product_type_counts(rows),
        "in_stock": dict(in_stock),
        "fragrance_slots": dict(fragrance_slots),
    }


def audit_curated_v2_workbook(workbook_path: str) -> dict[str, Any]:
    workbook = Path(workbook_path).resolve()
    if not workbook.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook}")

    wb = load_workbook(workbook)
    if CURATED_WORKBOOK_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook {workbook} has no '{CURATED_WORKBOOK_SHEET}' sheet.")
    ws = wb[CURATED_WORKBOOK_SHEET]

    headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    missing_headers = sorted(REQUIRED_HEADERS - set(headers))

    normalized_rows: list[dict[str, Any]] = []
    issues: list[WorkbookRowIssue] = []
    soft_fixes: list[dict[str, Any]] = []
    duplicate_tracker: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    fragrance_rows = 0
    fragrance_attrs = Counter()

    for sheet_row in range(2, ws.max_row + 1):
        raw_row = {
            headers[col_idx - 1]: ws.cell(row=sheet_row, column=col_idx).value
            for col_idx in range(1, ws.max_column + 1)
        }
        row: dict[str, Any] = {}
        row_issues: list[WorkbookRowIssue] = []
        for key, value in raw_row.items():
            if key in ALL_JSON_FIELDS:
                row[key] = _json_load_cell(value, field_name=key, sheet_row=sheet_row, issues=row_issues)
            else:
                row[key] = value

        for field_name in CRITICAL_ROW_FIELDS:
            if row.get(field_name) in (None, ""):
                row_issues.append(
                    WorkbookRowIssue(sheet_row=sheet_row, field=field_name, reason="Missing critical field.")
                )

        row["name"] = str(row.get("name") or "").strip()
        row["brand"] = str(row.get("brand") or "").strip()
        row["currency"] = str(row.get("currency") or "KZT").strip() or "KZT"
        row["category"] = str(row.get("category") or "").strip()
        row["product_type"] = str(row.get("product_type") or "").strip()
        row["step"] = str(row.get("step") or "").strip()
        row["image_url"] = str(row.get("image_url") or "").strip()
        row["description"] = str(row.get("description") or "").strip()
        row["application_text"] = str(row.get("application_text") or "").strip()
        row["ingredients_inci"] = str(row.get("ingredients_inci") or "").strip()
        row["volume_raw"] = str(row.get("volume_raw") or "").strip()
        row["in_stock"] = _normalize_bool(row.get("in_stock"))
        row["price"] = _normalize_price(row.get("price"), sheet_row=sheet_row, issues=row_issues)
        row["source_row"] = sheet_row

        source_product_id = str(raw_row.get("source_product_id") or "").strip()
        if not source_product_id:
            source_product_id = _generated_source_product_id(row, sheet_row=sheet_row)
            soft_fixes.append(
                {
                    "sheet_row": sheet_row,
                    "field": "source_product_id",
                    "reason": "Generated stable source_product_id for curated added row.",
                    "generated_value": source_product_id,
                }
            )
        if len(source_product_id) > 60:
            digest = hashlib.sha1(source_product_id.encode("utf-8")).hexdigest()[:12]
            shortened = f"cv2:{sheet_row}:{digest}"
            soft_fixes.append(
                {
                    "sheet_row": sheet_row,
                    "field": "source_product_id",
                    "reason": "Shortened source_product_id to fit Product.source_product_id length.",
                    "generated_value": shortened,
                }
            )
            source_product_id = shortened
        row["source_product_id_raw"] = source_product_id

        allowed_types = CURATED_V2_CANONICAL_TYPES.get(row["category"])
        if not allowed_types:
            row_issues.append(
                WorkbookRowIssue(
                    sheet_row=sheet_row,
                    field="category",
                    reason="Category is outside the supported ontology.",
                    raw_value=row["category"],
                )
            )
        elif row["product_type"] not in allowed_types:
            row_issues.append(
                WorkbookRowIssue(
                    sheet_row=sheet_row,
                    field="product_type",
                    reason="product_type is outside the supported ontology for category.",
                    raw_value=row["product_type"],
                )
            )

        if row["category"] == "fragrance":
            fragrance_rows += 1
            attrs = row.get("attrs") or {}
            if attrs.get("scent_family"):
                fragrance_attrs["scent_family"] += 1
            else:
                row_issues.append(
                    WorkbookRowIssue(
                        sheet_row=sheet_row,
                        field="attrs.scent_family",
                        reason="Fragrance row is missing scent_family.",
                    )
                )
            if attrs.get("notes"):
                fragrance_attrs["notes"] += 1
            else:
                row_issues.append(
                    WorkbookRowIssue(
                        sheet_row=sheet_row,
                        field="attrs.notes",
                        reason="Fragrance row is missing notes.",
                    )
                )
            if attrs.get("intensity"):
                fragrance_attrs["intensity"] += 1
            else:
                row_issues.append(
                    WorkbookRowIssue(
                        sheet_row=sheet_row,
                        field="attrs.intensity",
                        reason="Fragrance row is missing intensity.",
                    )
                )

        duplicate_key = (row["name"].lower(), row["brand"].lower(), row["product_type"].lower())
        duplicate_tracker[duplicate_key].append(
            {
                "sheet_row": sheet_row,
                "source_product_id": row["source_product_id_raw"],
                "volume_raw": row["volume_raw"],
            }
        )

        issues.extend(row_issues)
        normalized_rows.append(row)

    duplicate_groups = []
    for (name, brand, product_type), members in duplicate_tracker.items():
        if len(members) < 2:
            continue
        duplicate_groups.append(
            {
                "name": name,
                "brand": brand,
                "product_type": product_type,
                "count": len(members),
                "rows": members,
            }
        )
    duplicate_groups.sort(key=lambda item: (-int(item["count"]), item["brand"], item["name"]))

    blocking_issues = [issue.as_dict() for issue in issues if issue.severity == "blocking"]
    coverage = coverage_from_rows(normalized_rows)
    ready = not missing_headers and not blocking_issues
    return {
        "workbook_path": str(workbook),
        "sheet": CURATED_WORKBOOK_SHEET,
        "rows_total": len(normalized_rows),
        "rows_valid": len(normalized_rows) if ready else len(normalized_rows) - len(blocking_issues),
        "required_headers_missing": missing_headers,
        "soft_fixes": soft_fixes,
        "soft_fix_count": len(soft_fixes),
        "blocking_issues": blocking_issues,
        "invalid_rows": blocking_issues,
        "duplicate_groups_count": len(duplicate_groups),
        "duplicate_groups": duplicate_groups,
        "fragrance_attrs_coverage": {
            "rows": fragrance_rows,
            "scent_family": int(fragrance_attrs.get("scent_family", 0)),
            "notes": int(fragrance_attrs.get("notes", 0)),
            "intensity": int(fragrance_attrs.get("intensity", 0)),
        },
        "coverage": coverage,
        "normalized_rows": normalized_rows,
        "ready_to_import": ready,
    }


def build_runtime_import_report_md(report: dict[str, Any]) -> str:
    audit = report.get("audit") or {}
    coverage = (report.get("import") or {}).get("coverage_after") or (audit.get("coverage") or {})
    smoke = report.get("smoke") or {}
    failures = smoke.get("failures") or []
    lines = [
        "# Curated V2 Runtime Import Report",
        "",
        "## Audit",
        f"- workbook: `{audit.get('workbook_path')}`",
        f"- rows total: **{audit.get('rows_total', 0)}**",
        f"- rows valid: **{audit.get('rows_valid', 0)}**",
        f"- soft fixes in pipeline: **{audit.get('soft_fix_count', 0)}**",
        f"- duplicate groups: **{audit.get('duplicate_groups_count', 0)}**",
        f"- blocking invalid rows: **{len(audit.get('blocking_issues') or [])}**",
        f"- fragrance attrs coverage: `{json.dumps(audit.get('fragrance_attrs_coverage') or {}, ensure_ascii=False, default=json_default)}`",
        "",
        "## Import",
        f"- executed: **{bool(report.get('executed'))}**",
        f"- products before: **{(report.get('import') or {}).get('products_before', 0)}**",
        f"- products after: **{(report.get('import') or {}).get('products_after', 0)}**",
        f"- by category: `{json.dumps(coverage.get('by_category') or {}, ensure_ascii=False, default=json_default)}`",
        f"- canonical coverage: `{json.dumps(coverage.get('canonical_coverage') or {}, ensure_ascii=False, default=json_default)}`",
        f"- in_stock: `{json.dumps(coverage.get('in_stock') or {}, ensure_ascii=False, default=json_default)}`",
        f"- fragrance slots: `{json.dumps(coverage.get('fragrance_slots') or {}, ensure_ascii=False, default=json_default)}`",
        "",
        "## Smoke",
        f"- catalog endpoint checks: **{len((smoke.get('catalog_endpoints') or {}))}**",
        f"- roadmap endpoint checks: **{len((smoke.get('roadmap_endpoints') or {}))}**",
        f"- checkout scenarios: **{len((smoke.get('checkout_scenarios') or {}))}**",
        f"- failures: **{len(failures)}**",
    ]
    if failures:
        lines.extend(["", "### Failing Checks"])
        lines.extend(f"- {item}" for item in failures)
    return "\n".join(lines) + "\n"


def build_final_verdict_md(report: dict[str, Any]) -> str:
    verdict = report.get("verdict") or {}
    blockers = verdict.get("blocking_issues") or []
    dataset_notes = verdict.get("dataset_rebuild_notes") or []
    lines = [
        "# Curated V2 Runtime Verdict",
        "",
        f"- safe_for_demo_catalog = **{'yes' if verdict.get('safe_for_demo_catalog') else 'no'}**",
        f"- safe_for_runtime_catalog = **{'yes' if verdict.get('safe_for_runtime_catalog') else 'no'}**",
        f"- safe_for_dataset_rebuild = **{'yes' if verdict.get('safe_for_dataset_rebuild') else 'no'}**",
        "",
        "## Blockers",
    ]
    if blockers:
        lines.extend(f"- {item}" for item in blockers)
    else:
        lines.append("- none")
    if dataset_notes:
        lines.extend(["", "## Dataset Rebuild Notes"])
        lines.extend(f"- {item}" for item in dataset_notes)
    return "\n".join(lines) + "\n"
