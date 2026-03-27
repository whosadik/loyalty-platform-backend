from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import CommandError
from django.db.models import Q

from catalog.management.commands.import_products_xlsx import (
    _clip,
    _norm_attrs,
    _norm_brand,
    _norm_category,
    _norm_list,
    _norm_product_type,
    _norm_step,
    _norm_strength,
    _norm_supported_skin_types,
    _normalize_product_images,
    _parse_json,
    _resolve_path,
    _str,
    _to_bool,
    _to_decimal,
)
from catalog.models import Product
from offers.models import OfferAssignment
from recs_analytics.models import RecommendationEvent
from roadmap_app.fragrance_slots import slot_of_fragrance
from roadmap_app.models import RoadmapPlan, RoadmapStep
from transactions.models import OwnedProduct, TransactionItem

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    load_workbook = None
    OPENPYXL_IMPORT_ERROR = exc
else:
    OPENPYXL_IMPORT_ERROR = None


CANONICAL_PRODUCT_TYPES: dict[str, list[str]] = {
    Product.Category.HAIRCARE: [
        "shampoo",
        "conditioner",
        "hair_mask",
        "hair_oil",
        "scalp_serum",
        "leave_in",
    ],
    Product.Category.SKINCARE: [
        "cleanser",
        "serum",
        "moisturizer",
        "spf",
        "toner",
        "mask",
        "eye_cream",
        "essence",
    ],
    Product.Category.MAKEUP: [
        "foundation",
        "mascara",
        "blush",
        "lipstick",
        "eyeshadow",
        "primer",
        "setting_spray",
    ],
    Product.Category.FRAGRANCE: [
        "edp",
        "edt",
        "body_mist",
        "perfume_oil",
    ],
}

FRAGRANCE_SLOT_FIELDS = ("scent_family", "notes", "intensity")
HAIRCARE_ATTR_FIELDS = ("hair_type", "scalp_type", "hair_thickness")
SKINCARE_ATTR_FIELDS = ("spf", "volume_ml", "area")
MAKEUP_ATTR_FIELDS = ("finish", "effect", "coverage", "undertone", "tone_family", "volume")

REQUIRED_COLUMNS = {"name", "brand", "price", "category", "product_type"}
PREFERRED_COLUMNS = {
    "id",
    "currency",
    "concerns",
    "attrs",
    "actives",
    "flags",
    "supported_skin_types",
    "strength",
    "in_stock",
    "step",
    "photo_url_primary",
    "photo_urls",
    "application_text",
    "ingredients_inci",
    "description_text",
    "product_type_raw",
    "concerns_raw",
    "supported_skin_types_raw",
    "area_raw",
    "volume_raw",
}

PRODUCT_FIELD_SOURCES: dict[str, list[str]] = {
    "source_product_id": ["id"],
    "name": ["name", "Name"],
    "brand": ["brand", "бренд", "Name4"],
    "price": ["price", "Price", "Price6"],
    "currency": ["currency"],
    "category": ["category"],
    "product_type": ["product_type", "product_type_raw", "Type"],
    "concerns": ["concerns", "concerns_raw"],
    "attrs": ["attrs"],
    "actives": ["actives"],
    "flags": ["flags"],
    "supported_skin_types": ["supported_skin_types", "supported_skin_types_raw"],
    "strength": ["strength"],
    "in_stock": ["in_stock"],
    "step": ["step"],
    "description": ["description_text", "описаниебренда", "Info1"],
    "application_text": ["application_text", "применение"],
    "ingredients_inci": ["ingredients_inci", "сосстав"],
    "volume_raw": ["volume_raw"],
    "image_url": ["photo_url_primary", "Image", "Image5"],
    "image_urls": ["photo_urls"],
    "raw_meta": [
        "product_type_raw",
        "concerns_raw",
        "supported_skin_types_raw",
        "area_raw",
        "galinkbase_1x27n_109_URL",
        "страна",
        "View",
        "View2",
        "Info",
        "Info1",
        "Info3",
    ],
}


@dataclass(frozen=True)
class RawRule:
    category: str
    product_type: str
    patterns: tuple[str, ...]


RAW_MAPPING_RULES: tuple[RawRule, ...] = (
    RawRule("fragrance", "edp", (r"парфюмерн\w+\s+вода", r"eau de parfum", r"\bedp\b")),
    RawRule("fragrance", "edt", (r"туалетн\w+\s+вода", r"eau de toilette", r"\bedt\b")),
    RawRule(
        "fragrance",
        "body_mist",
        (r"body mist", r"hair mist", r"парфюмир\w+\s+мист", r"мист для тела"),
    ),
    RawRule("fragrance", "perfume_oil", (r"perfume oil", r"маслян\w+\s+дух", r"духи[- ]масло")),
    RawRule("haircare", "shampoo", (r"шампун", r"\bshampoo\b")),
    RawRule(
        "haircare",
        "conditioner",
        (r"кондиционер", r"бальзам\w*\s+для\s+волос", r"бальзамы\s+и\s+кондиционеры", r"\bconditioner\b"),
    ),
    RawRule("haircare", "hair_mask", (r"маск\w*\s+для\s+волос", r"hair mask")),
    RawRule("haircare", "hair_oil", (r"масл\w*\s+для\s+волос", r"hair oil")),
    RawRule(
        "haircare",
        "scalp_serum",
        (r"(сыворотк\w*|serum).*(кож\w+\s+голов|scalp)", r"scalp serum"),
    ),
    RawRule("haircare", "leave_in", (r"несмыва", r"leave[- ]?in")),
    RawRule("skincare", "spf", (r"\bspf\b", r"sunscreen", r"sun cream", r"sun stick", r"\buv\b")),
    RawRule("skincare", "eye_cream", (r"крем\w*.*глаз", r"eye cream")),
    RawRule("skincare", "toner", (r"тонер", r"тоник", r"\btoner\b")),
    RawRule("skincare", "essence", (r"эссенц", r"\bessence\b")),
    RawRule("skincare", "mask", (r"маск\w*\s+для\s+лиц", r"mask", r"скраб")),
    RawRule("skincare", "serum", (r"сыворотк\w*.*лиц", r"\bserum\b", r"\bampoule\b")),
    RawRule(
        "skincare",
        "cleanser",
        (
            r"очищающ",
            r"умыван",
            r"пенк\w*.*лиц",
            r"мицел",
            r"гидрофильн\w*\s+(масл|бальзам)",
            r"\bcleanser\b",
            r"\bcleansing\b",
        ),
    ),
    RawRule(
        "skincare",
        "moisturizer",
        (
            r"крем\w*.*лиц",
            r"эмульси\w*.*лиц",
            r"флюид\w*.*лиц",
            r"\bmoistur",
            r"крем-гель\s+для\s+лица",
        ),
    ),
    RawRule("makeup", "foundation", (r"тональн", r"\bfoundation\b", r"skin tint")),
    RawRule("makeup", "mascara", (r"туш", r"\bmascara\b")),
    RawRule("makeup", "blush", (r"румян", r"\bblush\b")),
    RawRule(
        "makeup",
        "lipstick",
        (r"губн\w+\s+помад", r"\blipstick\b", r"lip cloud", r"lip balm", r"lip gloss"),
    ),
    RawRule("makeup", "eyeshadow", (r"тен\w+\s+для\s+век", r"палетк\w+.*век", r"\beyeshadow\b")),
    RawRule("makeup", "primer", (r"праймер", r"\bprimer\b")),
    RawRule(
        "makeup",
        "setting_spray",
        (r"фиксирующ\w+\s+(спрей|мист)", r"setting spray", r"fixing spray"),
    ),
)

OUTSIDE_SCOPE_PATTERNS: tuple[tuple[str, str], ...] = (
    ("body_or_bath", r"(для тела|гель для душа|душа|ванн|bath|body lotion|body balm|body cream|body wash)"),
    ("hands_feet", r"(для рук|для стоп|для ног|hand cream|foot cream)"),
    ("household", r"(стирк|laundry|ополаскиватель для полости рта|mouthwash)"),
    ("tools_accessories", r"(кист[ья]|мочал|массажер|салфет|gift set|подарочн\w+\s+набор|набор)"),
    ("non_candidate_makeup", r"(пудр|powder|карандаш|eyeliner|brow|хайлайтер|highlighter|concealer)"),
    ("patches", r"(патч)"),
    ("face_mist", r"(face mist|мист.*для лица|спрей.*для лица)"),
)


def _norm_text(value: Any) -> str:
    s = _str(value).lower().replace("ё", "е")
    s = s.replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _json_default(value: Any, default: Any) -> Any:
    parsed = _parse_json(value, default)
    return default if parsed is None else parsed


def _load_workbook_rows(xlsx_path: str, sheet_name: str = "") -> dict[str, Any]:
    if load_workbook is None:
        raise CommandError(f"openpyxl is required: {OPENPYXL_IMPORT_ERROR}")

    path = _resolve_path(xlsx_path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        profiles: list[dict[str, Any]] = []
        chosen_title = ""
        best_score = -1
        for ws in wb.worksheets:
            headers: list[str] = []
            iterator = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(iterator, None)
            if header_row is not None:
                headers = [str(x).strip() if x is not None else "" for x in header_row]
            score = len(REQUIRED_COLUMNS & set(headers)) * 10 + len(PREFERRED_COLUMNS & set(headers))
            profiles.append(
                {
                    "sheet_name": ws.title,
                    "row_count": max(int(ws.max_row or 1) - 1, 0),
                    "column_count": len([x for x in headers if x]),
                    "columns": headers,
                    "catalog_score": score,
                }
            )
            if sheet_name:
                if ws.title == sheet_name:
                    chosen_title = ws.title
            elif score > best_score:
                chosen_title = ws.title
                best_score = score

        if sheet_name and sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        if not chosen_title:
            raise CommandError("Could not determine catalog sheet in workbook")

        ws = wb[chosen_title]
        iterator = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            raise CommandError("XLSX file is empty")
        headers = [str(x).strip() if x is not None else "" for x in header_row]
        rows: list[dict[str, Any]] = []
        for idx, raw_row in enumerate(iterator, start=2):
            row = {headers[col_idx]: raw_row[col_idx] for col_idx in range(len(headers)) if headers[col_idx]}
            row["__source_row__"] = idx
            rows.append(row)
        return {
            "path": path,
            "sheetnames": list(wb.sheetnames),
            "selected_sheet": chosen_title,
            "sheet_profiles": profiles,
            "columns": headers,
            "rows": rows,
        }
    finally:
        wb.close()


def _collect_sale_meta_from_row(row: dict[str, Any]) -> dict[str, Any]:
    sale_meta: dict[str, Any] = {}
    for key in ("original_price", "old_price", "price_old", "rrp", "compare_at_price"):
        value = _to_decimal(row.get(key))
        if value is not None:
            sale_meta["original_price"] = str(value)
            break
    for key in ("discount", "discount_percent", "sale_percent"):
        value = _to_decimal(row.get(key))
        if value is not None and value > 0:
            sale_meta["discount"] = int(value)
            break
    return sale_meta


def _collect_social_meta_from_row(row: dict[str, Any]) -> dict[str, Any]:
    social_meta: dict[str, Any] = {}
    for key in ("rating", "avg_rating"):
        value = _to_decimal(row.get(key))
        if value is not None:
            social_meta["rating"] = str(value)
            break
    for key in ("reviews_count", "reviews", "ratings_count"):
        value = _to_decimal(row.get(key))
        if value is not None and value >= 0:
            social_meta["reviews_count"] = int(value)
            break
    return social_meta


def build_normalized_product_payload(row: dict[str, Any]) -> dict[str, Any]:
    image_urls = _json_default(row.get("photo_urls"), [])
    if not isinstance(image_urls, list):
        image_urls = []
    image_urls = [_str(x) for x in image_urls if _str(x)]
    image_url, normalized_image_urls = _normalize_product_images(row.get("photo_url_primary"), image_urls)

    category = _norm_category(row.get("category"))
    product_type = _norm_product_type(row.get("product_type"))
    raw_meta = {
        "catalog_source": "goldapple_excel",
        "source_row": int(row.get("__source_row__") or 0),
        "source_sheet": _str(row.get("__sheet__")),
        "product_type_raw": _str(row.get("product_type_raw")),
        "concerns_raw": _str(row.get("concerns_raw")),
        "supported_skin_types_raw": _str(row.get("supported_skin_types_raw")),
        "area_raw": _str(row.get("area_raw")),
        "raw_name": _str(row.get("Name")),
        "raw_brand": _str(row.get("бренд")) or _str(row.get("Name4")),
        "raw_type_title": _str(row.get("Type")),
        "raw_url": _str(row.get("galinkbase_1x27n_109_URL")),
        "country_raw": _str(row.get("страна")),
        **_collect_sale_meta_from_row(row),
        **_collect_social_meta_from_row(row),
    }
    raw_meta = {
        k: v
        for k, v in raw_meta.items()
        if v is not None and v != "" and v != [] and v != {}
    }

    return {
        "source_product_id": _clip(_str(row.get("id")), 64),
        "name": _clip(_str(row.get("name") or row.get("Name")), 200),
        "brand": _clip(_norm_brand(row.get("brand") or row.get("бренд") or row.get("Name4")), 120),
        "price": _to_decimal(row.get("price") or row.get("Price") or row.get("Price6")),
        "currency": _clip(_str(row.get("currency")), 8),
        "category": category,
        "product_type": _clip(product_type, 50),
        "concerns": _norm_list(row.get("concerns")),
        "attrs": _norm_attrs(row.get("attrs")),
        "actives": _norm_list(row.get("actives")),
        "flags": _norm_list(row.get("flags")),
        "supported_skin_types": _norm_supported_skin_types(row.get("supported_skin_types")),
        "strength": _clip(_norm_strength(row.get("strength")), 20),
        "in_stock": _to_bool(row.get("in_stock"), default=True),
        "step": _norm_step(row.get("step"), category=category, product_type=product_type),
        "image_url": image_url,
        "image_urls": normalized_image_urls,
        "description": _str(row.get("description_text") or row.get("описаниебренда")),
        "application_text": _str(row.get("application_text") or row.get("применение")),
        "ingredients_inci": _str(row.get("ingredients_inci") or row.get("сосстав")),
        "volume_raw": _str(row.get("volume_raw")),
        "raw_meta": raw_meta,
    }


def _coalesce_raw_text(row: dict[str, Any]) -> str:
    parts = [
        row.get("product_type_raw"),
        row.get("Type"),
        row.get("name"),
        row.get("Name"),
        row.get("description_text"),
        row.get("Info"),
        row.get("Info1"),
        row.get("Info3"),
    ]
    return " ".join(_norm_text(x) for x in parts if _str(x))


def _outside_scope_reasons(text: str) -> list[str]:
    reasons: list[str] = []
    for name, pattern in OUTSIDE_SCOPE_PATTERNS:
        if re.search(pattern, text):
            reasons.append(name)
    return reasons


def _match_raw_rules(text: str, attrs: dict[str, Any]) -> list[tuple[str, str, str]]:
    matches: list[tuple[str, str, str]] = []
    for rule in RAW_MAPPING_RULES:
        for pattern in rule.patterns:
            if re.search(pattern, text):
                matches.append((rule.category, rule.product_type, pattern))
                break
    if attrs.get("spf"):
        matches.append(("skincare", "spf", "attrs.spf"))
    return matches


def classify_row_mapping(row: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    text = _coalesce_raw_text(row)
    attrs = payload.get("attrs") or {}
    matches = _match_raw_rules(text, attrs)
    unique_matches = sorted({(category, product_type) for category, product_type, _ in matches})
    outside_reasons = _outside_scope_reasons(text)
    normalized_pair = (payload["category"], payload["product_type"])

    if "face_mist" in outside_reasons and normalized_pair == ("fragrance", "body_mist"):
        return {
            "status": "ambiguous",
            "expected_category": "",
            "expected_product_type": "",
            "reason": "face_mist_looks_skincare_not_fragrance",
            "text_excerpt": text[:180],
        }

    if len(unique_matches) == 1:
        expected_category, expected_product_type = unique_matches[0]
        status = "confident"
        reason = "raw_type_consistent"
        if normalized_pair != (expected_category, expected_product_type):
            status = "ambiguous"
            reason = "normalized_mismatch_vs_raw"
        return {
            "status": status,
            "expected_category": expected_category,
            "expected_product_type": expected_product_type,
            "reason": reason,
            "text_excerpt": text[:180],
        }

    if len(unique_matches) > 1:
        if ("skincare", "spf") in unique_matches:
            return {
                "status": "confident",
                "expected_category": "skincare",
                "expected_product_type": "spf",
                "reason": "raw_type_spf_priority",
                "text_excerpt": text[:180],
            }
        if ("skincare", "cleanser") in unique_matches and ("skincare", "moisturizer") in unique_matches:
            if "очища" in text or "умыв" in text or "cleansing" in text:
                return {
                    "status": "confident",
                    "expected_category": "skincare",
                    "expected_product_type": "cleanser",
                    "reason": "cleanser_disambiguated_from_text",
                    "text_excerpt": text[:180],
                }
        return {
            "status": "ambiguous",
            "expected_category": "",
            "expected_product_type": "",
            "reason": "multiple_raw_matches",
            "text_excerpt": text[:180],
        }

    if outside_reasons:
        return {
            "status": "unmapped",
            "expected_category": "",
            "expected_product_type": "",
            "reason": "outside_project_ontology:" + ",".join(sorted(outside_reasons)),
            "text_excerpt": text[:180],
        }

    if payload["category"] in CANONICAL_PRODUCT_TYPES and payload["product_type"] in CANONICAL_PRODUCT_TYPES[payload["category"]]:
        return {
            "status": "ambiguous",
            "expected_category": payload["category"],
            "expected_product_type": payload["product_type"],
            "reason": "normalized_only_no_strong_raw_support",
            "text_excerpt": text[:180],
        }

    return {
        "status": "unmapped",
        "expected_category": "",
        "expected_product_type": "",
        "reason": "cannot_map_to_supported_ontology",
        "text_excerpt": text[:180],
    }


def _field_fill_rate(rows: list[dict[str, Any]], field_name: str) -> dict[str, Any]:
    if not rows:
        return {"non_empty": 0, "pct": 0.0}
    non_empty = 0
    for row in rows:
        value = row.get(field_name)
        if isinstance(value, str):
            if value.strip():
                non_empty += 1
        elif isinstance(value, (list, dict, tuple, set)):
            if value:
                non_empty += 1
        elif value is not None:
            non_empty += 1
    return {"non_empty": non_empty, "pct": round(non_empty / len(rows) * 100.0, 2)}


def _product_field_mapping_report(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    report: list[dict[str, Any]] = []
    col_set = set(columns)
    for product_field, source_columns in PRODUCT_FIELD_SOURCES.items():
        present = [column for column in source_columns if column in col_set]
        usable = any(_field_fill_rate(rows, column)["non_empty"] > 0 for column in present)
        notes = ""
        if product_field == "raw_meta" and "raw_meta" not in col_set:
            notes = "derived from source/raw columns; no direct raw_meta column in workbook"
        elif not present:
            notes = "missing direct source column"
        report.append(
            {
                "product_field": product_field,
                "source_columns": present,
                "status": "mapped" if usable else "missing",
                "notes": notes,
            }
        )
    return report


def _normalized_rows(rows: list[dict[str, Any]], sheet_name: str) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        row = dict(row)
        row["__sheet__"] = sheet_name
        normalized.append(build_normalized_product_payload(row))
    return normalized


def _category_product_type_coverage(normalized_rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_category_type = Counter((row["category"], row["product_type"]) for row in normalized_rows if row["name"])
    categories: dict[str, Any] = {}
    for category, vocab in CANONICAL_PRODUCT_TYPES.items():
        counts = {product_type: int(by_category_type.get((category, product_type), 0)) for product_type in vocab}
        categories[category] = {
            "total_rows": int(sum(counts.values())),
            "counts": counts,
            "missing": [product_type for product_type, count in counts.items() if count == 0],
        }
    return categories


def _attrs_quality(normalized_rows: list[dict[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in normalized_rows:
        grouped[row["category"]].append(row)

    def coverage(rows_by_category: list[dict[str, Any]], keys: tuple[str, ...]) -> dict[str, Any]:
        total = len(rows_by_category)
        out: dict[str, Any] = {}
        for key in keys:
            count = 0
            for row in rows_by_category:
                attrs = row.get("attrs") or {}
                value = attrs.get(key) if isinstance(attrs, dict) else None
                if value is not None and value != "" and value != [] and value != {}:
                    count += 1
            out[key] = {"non_empty": count, "pct": round(count / total * 100.0, 2) if total else 0.0}
        return out

    fragrance_rows = grouped.get("fragrance", [])
    slot_counts = Counter(
        slot_of_fragrance(row.get("attrs") or {}, raw_meta=row.get("raw_meta") or {})
        for row in fragrance_rows
        if row.get("product_type") in CANONICAL_PRODUCT_TYPES["fragrance"]
    )
    result["fragrance"] = {
        "total": len(fragrance_rows),
        "coverage": coverage(fragrance_rows, FRAGRANCE_SLOT_FIELDS),
        "slot_distribution": dict(slot_counts),
    }
    result["haircare"] = {
        "total": len(grouped.get("haircare", [])),
        "coverage": coverage(grouped.get("haircare", []), HAIRCARE_ATTR_FIELDS),
    }
    result["skincare"] = {
        "total": len(grouped.get("skincare", [])),
        "coverage": coverage(grouped.get("skincare", []), SKINCARE_ATTR_FIELDS),
    }
    result["makeup"] = {
        "total": len(grouped.get("makeup", [])),
        "coverage": coverage(grouped.get("makeup", []), MAKEUP_ATTR_FIELDS),
    }
    return result


def _mapping_quality_report(rows: list[dict[str, Any]], normalized_rows: list[dict[str, Any]]) -> dict[str, Any]:
    reason_counts = Counter()
    status_counts = Counter()
    examples: list[dict[str, Any]] = []
    confident_by_category = Counter()
    confident_by_type = Counter()

    for row, normalized in zip(rows, normalized_rows):
        mapping = classify_row_mapping(row, normalized)
        status_counts[mapping["status"]] += 1
        reason_counts[mapping["reason"]] += 1
        if mapping["status"] == "confident":
            confident_by_category[mapping["expected_category"]] += 1
            confident_by_type[(mapping["expected_category"], mapping["expected_product_type"])] += 1
        if mapping["status"] != "confident" and len(examples) < 25:
            examples.append(
                {
                    "source_row": row.get("__source_row__"),
                    "name": normalized.get("name"),
                    "brand": normalized.get("brand"),
                    "normalized_category": normalized.get("category"),
                    "normalized_product_type": normalized.get("product_type"),
                    "status": mapping["status"],
                    "reason": mapping["reason"],
                    "raw_type": _str(row.get("product_type_raw") or row.get("Type")),
                }
            )

    total = len(rows)
    ambiguous = int(status_counts.get("ambiguous", 0))
    unmapped = int(status_counts.get("unmapped", 0))
    suspicious_ratio = round((ambiguous + unmapped) / total * 100.0, 2) if total else 0.0
    return {
        "total_rows": total,
        "status_counts": dict(status_counts),
        "reason_counts": dict(reason_counts),
        "confident_by_category": dict(confident_by_category),
        "confident_by_type": {
            f"{category}:{product_type}": count
            for (category, product_type), count in sorted(confident_by_type.items())
        },
        "suspicious_ratio_pct": suspicious_ratio,
        "examples": examples,
    }


def _runtime_compatibility_report(
    normalized_rows: list[dict[str, Any]],
    mapping_quality: dict[str, Any],
    attr_quality: dict[str, Any],
) -> dict[str, Any]:
    issues: list[str] = []
    notes: dict[str, list[str]] = defaultdict(list)

    suspicious_ratio = float(mapping_quality.get("suspicious_ratio_pct") or 0.0)
    unmapped = int(mapping_quality.get("status_counts", {}).get("unmapped", 0))

    if suspicious_ratio >= 10.0 or unmapped >= 10:
        issues.append("semantic_noise_in_category_product_type_mapping")

    if attr_quality["fragrance"]["total"] > 0:
        if attr_quality["fragrance"]["coverage"]["scent_family"]["pct"] < 80.0:
            issues.append("fragrance_scent_family_coverage_too_low")
        if attr_quality["fragrance"]["coverage"]["notes"]["pct"] < 80.0:
            issues.append("fragrance_notes_coverage_too_low")
        if attr_quality["fragrance"]["coverage"]["intensity"]["pct"] < 80.0:
            issues.append("fragrance_intensity_coverage_too_low")

    coverage = _category_product_type_coverage(normalized_rows)
    for category in ("haircare", "skincare", "makeup"):
        missing = coverage[category]["missing"]
        if missing:
            notes["roadmap"].append(f"{category} candidate coverage has gaps: {', '.join(missing)}")

    if suspicious_ratio:
        notes["roadmap"].append("roadmap candidate pools would inherit noisy category/product_type labels")
        notes["recs"].append("recs/reranker feature space depends on category/product_type/attrs and would be polluted by mislabeled rows")
        notes["offers"].append("offer product_id/product_type targeting would inherit mislabeled catalog rows")
        notes["teacher_dataset"].append("teacher/planner datasets would learn from catalog ontology that contains semantic mismatches")

    slot_dist = attr_quality["fragrance"]["slot_distribution"]
    if attr_quality["fragrance"]["total"] > 0 and len(slot_dist) < 3:
        issues.append("fragrance_slot_distribution_collapse")

    if not issues:
        verdict = "fully_compatible"
    elif suspicious_ratio < 5.0 and unmapped == 0:
        verdict = "compatible_with_normalization_layer"
    else:
        verdict = "not_safe_yet"

    return {
        "verdict": verdict,
        "issues": issues,
        "subsystems": {
            "roadmap": notes["roadmap"],
            "recs": notes["recs"],
            "offers": notes["offers"],
            "fragrance_slot_logic": [
                "slot_of_fragrance can run on workbook attrs/raw_meta"
                if attr_quality["fragrance"]["total"]
                else "no fragrance rows detected"
            ],
            "teacher_dataset": notes["teacher_dataset"],
        },
    }


def _db_reference_audit() -> dict[str, Any]:
    products_total = Product.objects.count()
    src_prefix = Product.objects.filter(source_product_id__startswith="SRC").count()
    blank_source = Product.objects.filter(source_product_id="").count()
    return {
        "products_total": products_total,
        "referenced_distinct_product_ids": {
            "transaction_item": TransactionItem.objects.exclude(product_id__isnull=True).values("product_id").distinct().count(),
            "owned_product": OwnedProduct.objects.exclude(product_id__isnull=True).values("product_id").distinct().count(),
            "recommendation_event": RecommendationEvent.objects.exclude(product_id__isnull=True).values("product_id").distinct().count(),
            "roadmap_step_recommended_product": RoadmapStep.objects.exclude(recommended_product_id__isnull=True).values("recommended_product_id").distinct().count(),
        },
        "reference_rows": {
            "transaction_item": TransactionItem.objects.exclude(product_id__isnull=True).count(),
            "owned_product": OwnedProduct.objects.exclude(product_id__isnull=True).count(),
            "recommendation_event": RecommendationEvent.objects.exclude(product_id__isnull=True).count(),
            "roadmap_step_recommended_product": RoadmapStep.objects.exclude(recommended_product_id__isnull=True).count(),
            "offer_assignment_scope_product_id": OfferAssignment.objects.filter(target__scope="product_id").count(),
            "offer_assignment_scope_product_type": OfferAssignment.objects.filter(target__scope="product_type").count(),
        },
        "protect_delete_blockers": {
            "transaction_item_fk": True,
            "owned_product_fk": True,
        },
        "set_null_or_cascade_risks": {
            "roadmap_step_recommended_product_set_null": True,
            "recommendation_event_product_cascade": True,
        },
        "synthetic_detection": {
            "source_product_id_src_prefix": src_prefix,
            "source_product_id_blank": blank_source,
            "source_product_id_other": products_total - src_prefix - blank_source,
        },
        "active_roadmap": {
            "plans": RoadmapPlan.objects.filter(is_active=True).count(),
            "recommended_steps": RoadmapStep.objects.filter(
                plan__is_active=True,
                recommended_product_id__isnull=False,
            ).count(),
        },
    }


def choose_migration_strategy(audit: dict[str, Any]) -> dict[str, Any]:
    runtime_verdict = audit["runtime_compatibility"]["verdict"]
    db_audit = audit["db_reference_audit"]

    reasons: list[str] = []
    blockers: list[str] = []
    warnings: list[str] = []
    strategy = "archive_add"
    hard_replace_blocked = (
        db_audit["referenced_distinct_product_ids"]["transaction_item"] > 0
        or db_audit["referenced_distinct_product_ids"]["owned_product"] > 0
    )

    if hard_replace_blocked:
        reasons.append("hard_replace blocked by PROTECT references from transactions and owned products")
    else:
        reasons.append("hard_replace not blocked by PROTECT references")

    if runtime_verdict == "not_safe_yet":
        blockers.append("workbook semantic quality is not safe for import")

    if hard_replace_blocked:
        blockers.append("catalog public read paths use Product.objects.all() without hiding archived legacy rows")
        warnings.append("archive_add would keep legacy products visible in /api/products and brand aggregates unless runtime read path changes")
    warnings.append("hard_replace would require clearing historical demo/runtime data across transactions, offers, recs, and roadmap references")

    if blockers:
        chosen = "blocked"
    elif hard_replace_blocked:
        chosen = "archive_add"
    else:
        chosen = "hard_replace"

    if chosen == "blocked":
        strategy = "fresh_rebuild"
        reasons.append("full replacement without runtime changes is only clean in a fresh/demo rebuild path")
    else:
        strategy = chosen

    return {
        "chosen_strategy": strategy,
        "status": chosen,
        "reasons": reasons,
        "blockers": blockers,
        "warnings": warnings,
    }


def audit_goldapple_catalog(xlsx_path: str, sheet_name: str = "") -> dict[str, Any]:
    workbook = _load_workbook_rows(xlsx_path, sheet_name=sheet_name)
    rows = workbook["rows"]
    normalized_rows = _normalized_rows(rows, workbook["selected_sheet"])

    field_mapping = _product_field_mapping_report(rows, workbook["columns"])
    mapping_quality = _mapping_quality_report(rows, normalized_rows)
    coverage = _category_product_type_coverage(normalized_rows)
    attr_quality = _attrs_quality(normalized_rows)
    runtime = _runtime_compatibility_report(normalized_rows, mapping_quality, attr_quality)
    db_refs = _db_reference_audit()

    missing_required = sorted(REQUIRED_COLUMNS - set(workbook["columns"]))
    missing_preferred = sorted(PREFERRED_COLUMNS - set(workbook["columns"]))
    overall_verdict = "not_safe_yet" if missing_required else runtime["verdict"]

    audit = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_path": str(workbook["path"]),
        "sheetnames": workbook["sheetnames"],
        "selected_sheet": workbook["selected_sheet"],
        "sheet_profiles": workbook["sheet_profiles"],
        "rows_total": len(rows),
        "columns": workbook["columns"],
        "required_columns_missing": missing_required,
        "preferred_columns_missing": missing_preferred,
        "product_field_mapping": field_mapping,
        "column_fill_rate": {
            column: _field_fill_rate(rows, column) for column in workbook["columns"] if column
        },
        "mapping_quality": mapping_quality,
        "category_product_type_coverage": coverage,
        "attrs_quality": attr_quality,
        "runtime_compatibility": runtime,
        "db_reference_audit": db_refs,
        "overall_verdict": overall_verdict,
    }
    audit["migration_strategy"] = choose_migration_strategy(audit)
    return audit


def _md_table(headers: list[str], rows: list[list[Any]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows or [["-"] * len(headers)]:
        lines.append("| " + " | ".join(str(value) for value in row) + " |")
    return "\n".join(lines)


def render_audit_markdown(audit: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append("# Goldapple Catalog Audit")
    lines.append("")
    lines.append(f"- File: `{audit['xlsx_path']}`")
    lines.append(f"- Selected sheet: `{audit['selected_sheet']}`")
    lines.append(f"- Overall verdict: **{audit['overall_verdict']}**")
    lines.append(f"- Runtime compatibility verdict: **{audit['runtime_compatibility']['verdict']}**")
    lines.append(f"- Migration strategy: **{audit['migration_strategy']['chosen_strategy']}**")
    lines.append("")
    lines.append("## Workbook profile")
    lines.append(
        _md_table(
            ["sheet", "rows", "columns", "catalog_score"],
            [
                [
                    profile["sheet_name"],
                    profile["row_count"],
                    profile["column_count"],
                    profile["catalog_score"],
                ]
                for profile in audit["sheet_profiles"]
            ],
        )
    )
    lines.append("")
    lines.append(f"- Required columns missing: {audit['required_columns_missing'] or 'none'}")
    lines.append(f"- Preferred columns missing: {audit['preferred_columns_missing'] or 'none'}")
    lines.append("")
    lines.append("## Product field mapping")
    lines.append(
        _md_table(
            ["Product field", "Source columns", "Status", "Notes"],
            [
                [
                    row["product_field"],
                    ", ".join(row["source_columns"]) or "-",
                    row["status"],
                    row["notes"] or "-",
                ]
                for row in audit["product_field_mapping"]
            ],
        )
    )
    lines.append("")
    lines.append("## Mapping quality")
    mq = audit["mapping_quality"]
    lines.append(_md_table(["status", "count"], [[status, count] for status, count in sorted(mq["status_counts"].items())]))
    lines.append("")
    lines.append(f"- suspicious ratio: **{mq['suspicious_ratio_pct']}%**")
    lines.append(
        _md_table(
            ["reason", "count"],
            [[reason, count] for reason, count in sorted(mq["reason_counts"].items(), key=lambda kv: (-kv[1], kv[0]))],
        )
    )
    lines.append("")
    lines.append("### Suspicious examples")
    lines.append(
        _md_table(
            ["row", "brand", "name", "normalized", "status", "reason", "raw_type"],
            [
                [
                    row["source_row"],
                    row["brand"],
                    row["name"],
                    f"{row['normalized_category']} / {row['normalized_product_type']}",
                    row["status"],
                    row["reason"],
                    row["raw_type"],
                ]
                for row in mq["examples"][:15]
            ],
        )
    )
    lines.append("")
    lines.append("## Candidate coverage")
    for category, payload in audit["category_product_type_coverage"].items():
        lines.append(f"### {category}")
        lines.append(_md_table(["product_type", "count"], [[product_type, count] for product_type, count in payload["counts"].items()]))
        lines.append(f"- missing: {payload['missing'] or 'none'}")
        lines.append("")
    lines.append("## Attr quality")
    for category, payload in audit["attrs_quality"].items():
        lines.append(f"### {category}")
        lines.append(
            _md_table(
                ["attr", "non-empty", "coverage %"],
                [[key, stats["non_empty"], stats["pct"]] for key, stats in payload["coverage"].items()],
            )
        )
        if category == "fragrance":
            lines.append(f"- slot distribution: {json.dumps(payload['slot_distribution'], ensure_ascii=False, sort_keys=True)}")
        lines.append("")
    lines.append("## Runtime compatibility")
    lines.append(f"- verdict: **{audit['runtime_compatibility']['verdict']}**")
    lines.append(f"- issues: {audit['runtime_compatibility']['issues'] or 'none'}")
    for subsystem, notes in audit["runtime_compatibility"]["subsystems"].items():
        lines.append(f"- {subsystem}: {notes or ['ok']}")
    lines.append("")
    lines.append("## DB reference audit")
    db = audit["db_reference_audit"]
    lines.append(
        _md_table(
            ["metric", "value"],
            [
                ["products_total", db["products_total"]],
                ["tx_item_distinct_products", db["referenced_distinct_product_ids"]["transaction_item"]],
                ["owned_distinct_products", db["referenced_distinct_product_ids"]["owned_product"]],
                ["recommendation_event_distinct_products", db["referenced_distinct_product_ids"]["recommendation_event"]],
                ["roadmap_recommended_distinct_products", db["referenced_distinct_product_ids"]["roadmap_step_recommended_product"]],
                ["offer_assignment_scope_product_id", db["reference_rows"]["offer_assignment_scope_product_id"]],
                ["offer_assignment_scope_product_type", db["reference_rows"]["offer_assignment_scope_product_type"]],
                ["synthetic_source_src_prefix", db["synthetic_detection"]["source_product_id_src_prefix"]],
                ["synthetic_source_blank", db["synthetic_detection"]["source_product_id_blank"]],
            ],
        )
    )
    lines.append("")
    lines.append("## Migration strategy")
    ms = audit["migration_strategy"]
    lines.append(f"- chosen: **{ms['chosen_strategy']}**")
    lines.append(f"- blockers: {ms['blockers'] or 'none'}")
    lines.append(f"- warnings: {ms['warnings'] or 'none'}")
    lines.append(f"- reasons: {ms['reasons'] or 'none'}")
    lines.append("")
    return "\n".join(lines).strip() + "\n"


def json_default(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, Path):
        return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def export_products_backup(products_qs, backup_path: Path) -> int:
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "id",
        "source_product_id",
        "name",
        "brand",
        "price",
        "currency",
        "category",
        "product_type",
        "concerns",
        "attrs",
        "step",
        "actives",
        "flags",
        "supported_skin_types",
        "strength",
        "in_stock",
        "image_url",
        "image_urls",
        "description",
        "application_text",
        "ingredients_inci",
        "volume_raw",
        "raw_meta",
        "created_at",
        "updated_at",
    ]
    count = 0
    with backup_path.open("w", encoding="utf-8") as handle:
        for row in products_qs.values(*fields).iterator():
            handle.write(json.dumps(row, ensure_ascii=False, default=json_default) + "\n")
            count += 1
    return count


def identify_legacy_synthetic_products():
    return Product.objects.filter(Q(source_product_id__startswith="SRC") | Q(source_product_id=""))


def export_queryset_jsonl(queryset, backup_path: Path, *, fields: list[str] | None = None) -> int:
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    rows = queryset.values(*fields) if fields else queryset.values()
    with backup_path.open("w", encoding="utf-8") as handle:
        for row in rows.iterator():
            handle.write(json.dumps(row, ensure_ascii=False, default=json_default) + "\n")
            count += 1
    return count
