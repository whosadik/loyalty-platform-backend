from __future__ import annotations

import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models.deletion import ProtectedError

from catalog.models import Product

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    load_workbook = None
    OPENPYXL_IMPORT_ERROR = exc
else:
    OPENPYXL_IMPORT_ERROR = None


ALLOWED_CATEGORIES = {x[0] for x in Product.Category.choices}
ALLOWED_STRENGTH = {x[0] for x in Product.Strength.choices}
ALLOWED_STEPS = {x[0] for x in Product.Step.choices}


def _resolve_path(raw_path: str) -> Path:
    p = Path(raw_path)
    if p.is_absolute() and p.exists():
        return p

    candidates = [
        Path.cwd() / p,
        settings.BASE_DIR / p,
        settings.BASE_DIR.parent / p,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise CommandError(f"File not found: {raw_path}")


def _str(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def _clip(value: str, max_len: int) -> str:
    if len(value) <= max_len:
        return value
    return value[:max_len]


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = _str(value)
    if not s:
        return None
    s = s.replace(" ", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) != 0
    s = _str(value).lower()
    if not s:
        return default
    return s in {"1", "true", "yes", "y", "on"}


def _parse_json(value: Any, default: Any) -> Any:
    if value is None:
        return default
    if isinstance(value, (list, dict)):
        return value
    s = _str(value)
    if not s:
        return default
    try:
        return json.loads(s)
    except Exception:
        return default


def _norm_category(value: Any) -> str:
    s = _str(value).lower()
    if not s:
        return Product.Category.SKINCARE
    if s in ALLOWED_CATEGORIES:
        return s

    mapping = {
        "уход": Product.Category.SKINCARE,
        "skincare": Product.Category.SKINCARE,
        "hair": Product.Category.HAIRCARE,
        "волос": Product.Category.HAIRCARE,
        "makeup": Product.Category.MAKEUP,
        "макияж": Product.Category.MAKEUP,
        "fragrance": Product.Category.FRAGRANCE,
        "парфюм": Product.Category.FRAGRANCE,
        "аромат": Product.Category.FRAGRANCE,
    }
    for key, mapped in mapping.items():
        if key in s:
            return mapped
    return Product.Category.SKINCARE


def _norm_product_type(value: Any) -> str:
    s = _str(value).lower()
    if not s:
        return "unknown"
    s = s.replace("-", "_")
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "unknown"


def _norm_strength(value: Any) -> str:
    s = _str(value).lower()
    if s in ALLOWED_STRENGTH:
        return s
    return Product.Strength.LOW


def _norm_step(step_raw: Any, *, category: str, product_type: str) -> str:
    if category == Product.Category.SKINCARE and product_type in ALLOWED_STEPS:
        return product_type
    s = _str(step_raw).lower()
    if s in ALLOWED_STEPS:
        return s
    return ""


def _norm_list(value: Any) -> list[str]:
    parsed = _parse_json(value, [])
    if not isinstance(parsed, list):
        return []
    out = []
    for item in parsed:
        s = _str(item).lower()
        if s:
            out.append(s)
    return out


def _norm_supported_skin_types(value: Any) -> list[str]:
    vals = _norm_list(value)
    if not vals:
        return []
    all_aliases = {
        "all",
        "any",
        "all_skin_types",
        "for_all",
        "для_всех",
        "для_всех_типов_кожи",
    }
    normalized = []
    for v in vals:
        cleaned = re.sub(r"\s+", "_", v)
        if cleaned in all_aliases:
            return []
        normalized.append(cleaned)
    return normalized


def _norm_attrs(value: Any) -> dict[str, Any]:
    parsed = _parse_json(value, {})
    if isinstance(parsed, dict):
        return parsed
    return {}


class Command(BaseCommand):
    help = "Import products from XLSX file into catalog."

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default="data/catalog/goldapple_300_products.xlsx",
            help="Path to XLSX file with products.",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="",
            help="Sheet name. Empty means first sheet.",
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Delete existing products before import.",
        )
        parser.add_argument(
            "--reset-related",
            action="store_true",
            help="When used with --replace, also clear product-dependent data (transactions/owned/offers/recs).",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=0,
            help="Limit rows for testing. 0 means import all rows.",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError(f"openpyxl is required: {OPENPYXL_IMPORT_ERROR}")

        file_path = _resolve_path(options["path"])
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet_name = (options.get("sheet") or "").strip()
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise CommandError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
                ws = wb[sheet_name]
            else:
                ws = wb[wb.sheetnames[0]]

            rows = ws.iter_rows(min_row=1, values_only=True)
            try:
                headers = [str(x).strip() if x is not None else "" for x in next(rows)]
            except StopIteration:
                raise CommandError("XLSX file is empty")

            col = {name: idx for idx, name in enumerate(headers)}
            required = {"name", "brand", "price", "category", "product_type"}
            missing = [x for x in required if x not in col]
            if missing:
                raise CommandError(f"Missing required columns: {missing}")

            limit = int(options.get("limit") or 0)
            to_create: list[Product] = []
            created = 0
            updated = 0
            skipped = 0

            with transaction.atomic():
                if options.get("replace"):
                    try:
                        Product.objects.all().delete()
                    except ProtectedError:
                        if not options.get("reset_related"):
                            raise CommandError(
                                "Catalog replace is blocked by protected product references "
                                "(transactions or owned products). "
                                "Use --replace --reset-related to clear dependent demo data."
                            )
                        from offers.models import OfferAssignment, OfferEvent
                        from recs_analytics.models import RecommendationEvent
                        from transactions.models import OwnedProduct, Transaction

                        OfferEvent.objects.all().delete()
                        OfferAssignment.objects.all().delete()
                        RecommendationEvent.objects.all().delete()
                        OwnedProduct.objects.all().delete()
                        Transaction.objects.all().delete()
                        Product.objects.all().delete()

                for row_index, row in enumerate(rows, start=2):
                    if limit and len(to_create) + created + updated >= limit:
                        break

                    name = _str(row[col["name"]])
                    product_type = _norm_product_type(row[col["product_type"]])
                    category = _norm_category(row[col["category"]])
                    if not name:
                        skipped += 1
                        continue

                    source_product_id = _str(row[col["id"]]) if "id" in col else ""
                    attrs = _norm_attrs(row[col["attrs"]]) if "attrs" in col else {}
                    image_urls = _parse_json(row[col["photo_urls"]], []) if "photo_urls" in col else []
                    if not isinstance(image_urls, list):
                        image_urls = []
                    image_urls = [_str(x) for x in image_urls if _str(x)]

                    payload = {
                        "name": _clip(name, 200),
                        "brand": _clip(_str(row[col["brand"]]), 120),
                        "price": _to_decimal(row[col["price"]]),
                        "source_product_id": _clip(source_product_id, 64),
                        "currency": _clip(_str(row[col["currency"]]) if "currency" in col else "", 8),
                        "category": category,
                        "product_type": _clip(product_type, 50),
                        "concerns": _norm_list(row[col["concerns"]]) if "concerns" in col else [],
                        "attrs": attrs,
                        "actives": _norm_list(row[col["actives"]]) if "actives" in col else [],
                        "flags": _norm_list(row[col["flags"]]) if "flags" in col else [],
                        "supported_skin_types": (
                            _norm_supported_skin_types(row[col["supported_skin_types"]])
                            if "supported_skin_types" in col
                            else []
                        ),
                        "strength": _clip(
                            _norm_strength(row[col["strength"]]) if "strength" in col else Product.Strength.LOW,
                            20,
                        ),
                        "in_stock": _to_bool(row[col["in_stock"]], default=True) if "in_stock" in col else True,
                        "step": _norm_step(
                            row[col["step"]] if "step" in col else "",
                            category=category,
                            product_type=product_type,
                        ),
                        "image_url": _clip(_str(row[col["photo_url_primary"]]) if "photo_url_primary" in col else "", 500),
                        "image_urls": image_urls,
                        "description": _str(row[col["description_text"]]) if "description_text" in col else "",
                        "application_text": _str(row[col["application_text"]]) if "application_text" in col else "",
                        "ingredients_inci": _str(row[col["ingredients_inci"]]) if "ingredients_inci" in col else "",
                        "volume_raw": _str(row[col["volume_raw"]]) if "volume_raw" in col else "",
                        "raw_meta": {
                            "product_type_raw": _str(row[col["product_type_raw"]]) if "product_type_raw" in col else "",
                            "concerns_raw": _str(row[col["concerns_raw"]]) if "concerns_raw" in col else "",
                            "supported_skin_types_raw": (
                                _str(row[col["supported_skin_types_raw"]]) if "supported_skin_types_raw" in col else ""
                            ),
                            "area_raw": _str(row[col["area_raw"]]) if "area_raw" in col else "",
                            "source_row": row_index,
                        },
                    }

                    if options.get("replace"):
                        to_create.append(Product(**payload))
                        continue

                    if source_product_id:
                        _, was_created = Product.objects.update_or_create(
                            source_product_id=source_product_id,
                            defaults=payload,
                        )
                    else:
                        _, was_created = Product.objects.update_or_create(
                            name=name,
                            brand=payload["brand"],
                            category=category,
                            product_type=product_type,
                            defaults=payload,
                        )
                    if was_created:
                        created += 1
                    else:
                        updated += 1

                if options.get("replace") and to_create:
                    Product.objects.bulk_create(to_create, batch_size=500)
                    created += len(to_create)

            self.stdout.write(self.style.SUCCESS("Catalog import completed"))
            self.stdout.write(f"File: {file_path}")
            self.stdout.write(f"Sheet: {ws.title}")
            self.stdout.write(f"Created: {created}")
            self.stdout.write(f"Updated: {updated}")
            self.stdout.write(f"Skipped: {skipped}")
        finally:
            wb.close()
