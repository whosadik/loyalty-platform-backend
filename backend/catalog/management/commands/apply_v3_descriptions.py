"""Apply the v3 normalized catalog (clean description / application_text /
ingredients_inci / volume_raw / attrs.specs) on top of existing products.

This is the safe reverse of the v3 re-import: we keep all existing product
ids (so roadmaps, routines, owned products, transactions, image references
stay intact) and only rewrite the noisy text fields.

Matching strategy:
  1. By source_product_id, with or without the legacy `ga:` prefix.
  2. Fallback: by (name + brand + category + product_type), case-insensitive.

Fields rewritten:
  - description
  - application_text
  - ingredients_inci
  - volume_raw
  - attrs (merged: existing keys preserved, "specs" + new keys from v3 added)

Fields NEVER touched: image, image_url, image_urls, price, source_product_id,
brand, name, category, product_type, raw_meta — anything users / FKs depend on.

Run:
    python manage.py apply_v3_descriptions
    python manage.py apply_v3_descriptions --dry-run
    python manage.py apply_v3_descriptions --path data/catalog/goldapple_300_products_curated_v3.xlsx
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction as db_tx

from catalog.models import Product

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    load_workbook = None
    _OPENPYXL_ERROR: Exception | None = exc
else:
    _OPENPYXL_ERROR = None


WRITABLE_FIELDS = (
    "description",
    "application_text",
    "ingredients_inci",
    "volume_raw",
)


def _str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_json(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    s = _str(value)
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None


def _resolve_path(raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute() and p.exists():
        return p
    candidates = [Path.cwd() / p, settings.BASE_DIR / p, settings.BASE_DIR.parent / p]
    for c in candidates:
        if c.exists():
            return c
    raise CommandError(f"File not found: {raw}")


def _build_lookup_keys(row: dict[str, Any]) -> tuple[str, tuple[str, str, str, str]]:
    """Return (source_id, (name, brand, category, product_type)) for matching."""
    source_id = _str(row.get("source_product_id"))
    composite = (
        _str(row.get("name")).lower(),
        _str(row.get("brand")).lower(),
        _str(row.get("category")).lower(),
        _str(row.get("product_type")).lower(),
    )
    return source_id, composite


def _find_product_for_row(
    row: dict[str, Any],
    by_source_id: dict[str, Product],
    by_composite: dict[tuple[str, str, str, str], list[Product]],
) -> Product | None:
    source_id, composite = _build_lookup_keys(row)
    if source_id:
        # The dump's products were imported with a "ga:" prefix; the v3 file
        # has the bare SKU. Match both forms.
        if source_id in by_source_id:
            return by_source_id[source_id]
        prefixed = f"ga:{source_id}"
        if prefixed in by_source_id:
            return by_source_id[prefixed]
    candidates = by_composite.get(composite) or []
    if len(candidates) == 1:
        return candidates[0]
    return None


class Command(BaseCommand):
    help = "Apply v3 normalized description / specs onto existing products by source id (preserving ids)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            default="data/catalog/goldapple_300_products_curated_v3.xlsx",
            help="Path to the v3 XLSX file.",
        )
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument(
            "--limit",
            type=int,
            default=0,
            help="Stop after applying N updates (0 = all).",
        )

    def handle(self, *args, **opts):
        if load_workbook is None:
            raise CommandError(f"openpyxl is required: {_OPENPYXL_ERROR}")

        path = _resolve_path(opts["path"])
        self.stdout.write(f"v3 source: {path}")
        self.stdout.write(f"dry_run: {opts['dry_run']}")

        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        rows: list[dict[str, Any]] = []
        for raw in rows_iter:
            row = {h: v for h, v in zip(headers, raw) if h}
            if not _str(row.get("name")):
                continue
            rows.append(row)
        wb.close()
        self.stdout.write(f"v3 rows: {len(rows)}")

        products = list(Product.objects.all())
        by_source_id: dict[str, Product] = {}
        by_composite: dict[tuple[str, str, str, str], list[Product]] = {}
        for p in products:
            sid = (p.source_product_id or "").strip()
            if sid:
                by_source_id[sid] = p
                if sid.startswith("ga:"):
                    by_source_id[sid.split(":", 1)[1]] = p
            key = (
                (p.name or "").lower(),
                (p.brand or "").lower(),
                (p.category or "").lower(),
                (p.product_type or "").lower(),
            )
            by_composite.setdefault(key, []).append(p)
        self.stdout.write(f"DB products: {len(products)}")

        stats = {
            "matched_by_source": 0,
            "matched_by_composite": 0,
            "no_match": 0,
            "updated": 0,
            "unchanged": 0,
            "ambiguous_composite": 0,
        }
        unmatched_samples: list[str] = []
        limit = int(opts.get("limit") or 0)

        with db_tx.atomic():
            for row in rows:
                product = _find_product_for_row(row, by_source_id, by_composite)
                if product is None:
                    source_id, composite = _build_lookup_keys(row)
                    if composite in by_composite and len(by_composite[composite]) > 1:
                        stats["ambiguous_composite"] += 1
                    stats["no_match"] += 1
                    if len(unmatched_samples) < 12:
                        unmatched_samples.append(
                            f"  no_match: source_id={source_id!r} name={row.get('name')!r} brand={row.get('brand')!r}"
                        )
                    continue

                source_id, _ = _build_lookup_keys(row)
                if source_id and product.source_product_id in {source_id, f"ga:{source_id}"}:
                    stats["matched_by_source"] += 1
                else:
                    stats["matched_by_composite"] += 1

                changed = False
                for field in WRITABLE_FIELDS:
                    new_value = _str(row.get(field))
                    if not new_value:
                        continue
                    if getattr(product, field, "") != new_value:
                        setattr(product, field, new_value)
                        changed = True

                v3_attrs = _parse_json(row.get("attrs"))
                if isinstance(v3_attrs, dict):
                    existing_attrs: dict[str, Any]
                    if isinstance(product.attrs, dict):
                        existing_attrs = dict(product.attrs)
                    else:
                        existing_attrs = {}
                    merged = dict(existing_attrs)
                    for k, v in v3_attrs.items():
                        # Always overwrite specs (only present in v3); for
                        # other keys, prefer v3 only if absent in existing.
                        if k == "specs":
                            merged["specs"] = v
                        elif k not in merged:
                            merged[k] = v
                    if merged != existing_attrs:
                        product.attrs = merged
                        changed = True

                if changed:
                    stats["updated"] += 1
                    if not opts["dry_run"]:
                        product.save(update_fields=[*WRITABLE_FIELDS, "attrs", "updated_at"])
                    if limit and stats["updated"] >= limit:
                        break
                else:
                    stats["unchanged"] += 1

            if opts["dry_run"]:
                db_tx.set_rollback(True)

        self.stdout.write("")
        self.stdout.write("--- summary ---")
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if unmatched_samples:
            self.stdout.write("")
            self.stdout.write("unmatched samples:")
            for line in unmatched_samples:
                self.stdout.write(line)
